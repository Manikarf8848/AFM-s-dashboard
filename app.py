import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import hashlib
import io
import report_builder
import history_db

st.set_page_config(page_title="LCY3 AFM Dashboard", layout="wide", page_icon="📊")

if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = True
if "saved_active_hashes" not in st.session_state:
    st.session_state.saved_active_hashes = []
if "duplicate_warnings" not in st.session_state:
    st.session_state.duplicate_warnings = []

DM = st.session_state.dark_mode

_bg      = "#0f1117" if DM else "#ffffff"
_bg2     = "#1a1d27" if DM else "#f0f4ff"
_bg3     = "#22263a" if DM else "#e8eaf6"
_card    = "#1e2235" if DM else "#ffffff"
_text    = "#e8eaf6" if DM else "#1a237e"
_sub     = "#8892b0" if DM else "#666666"
_border  = "#3949ab"
_accent  = "#7986cb" if DM else "#3949ab"

_css_sidebar_border   = "#2a2d3e" if DM else "#e0e0e0"
# FIXED: Much higher contrast for light mode header
_css_header_bg        = "linear-gradient(135deg, #0d1333 0%, #1a237e 45%, #3949ab 100%)" if DM else "linear-gradient(135deg, #1a237e 0%, #283593 50%, #3949ab 100%)"
_css_header_shadow    = "0 4px 32px rgba(57,73,171,0.6), 0 0 0 1px rgba(121,134,203,0.15)" if DM else "0 4px 24px rgba(57,73,171,0.5), 0 0 0 1px rgba(57,73,171,0.3)"
_css_kpi_shadow       = "0 2px 20px rgba(0,0,0,0.4), 0 0 0 1px rgba(57,73,171,0.12)" if DM else "0 2px 16px rgba(0,0,0,0.08)"
_css_kpi_hover_shadow = "0 12px 36px rgba(57,73,171,0.55), 0 0 0 1px rgba(121,134,203,0.25)" if DM else "0 10px 30px rgba(57,73,171,0.25)"
_css_profile_shadow   = "0 4px 24px rgba(0,0,0,0.4), 0 0 0 1px rgba(57,73,171,0.15)" if DM else "0 2px 16px rgba(0,0,0,0.08)"
_css_profile_hover    = "0 8px 32px rgba(57,73,171,0.5), 0 0 0 1px rgba(121,134,203,0.2)" if DM else "0 6px 24px rgba(57,73,171,0.2)"
_css_tabs_hover_bg    = "rgba(57,73,171,0.1)" if DM else "rgba(57,73,171,0.06)"
_css_tabs_active_bg   = "rgba(57,73,171,0.12)" if DM else "rgba(57,73,171,0.07)"
_css_input_border     = "#3949ab" if DM else "#c5cae9"
_css_rc_bg1           = "#1e2235" if DM else "#fff3e0"
_css_rc_bg2           = "#22263a" if DM else "#fff8e1"
_css_rc_shadow        = "0 2px 16px rgba(0,0,0,0.3)" if DM else "0 2px 12px rgba(0,0,0,0.06)"
_css_top_bg           = "linear-gradient(135deg, #1e2235, #252847)" if DM else "linear-gradient(135deg, #fffde7, #fff8e1)"
_css_top_border       = "2px solid rgba(245,158,11,0.4)" if DM else "2px solid rgba(245,158,11,0.5)"
_css_top_shadow       = "0 4px 24px rgba(245,158,11,0.15), 0 0 0 1px rgba(245,158,11,0.1)" if DM else "0 4px 16px rgba(245,158,11,0.2)"
_css_top_hover        = "0 8px 32px rgba(245,158,11,0.25)" if DM else "0 8px 24px rgba(245,158,11,0.3)"
_css_chart_shadow     = "0 2px 20px rgba(0,0,0,0.35), 0 0 0 1px rgba(57,73,171,0.1)" if DM else "0 2px 14px rgba(0,0,0,0.07)"
_css_chart_hover      = "0 8px 32px rgba(57,73,171,0.35), 0 0 0 1px rgba(121,134,203,0.2)" if DM else "0 6px 24px rgba(57,73,171,0.18)"
_css_btn_bg           = "linear-gradient(135deg, #1a237e, #3949ab)" if DM else "linear-gradient(135deg, #3949ab, #5c6bc0)"
_css_uploader_bg      = "rgba(30,34,53,0.7)" if DM else "rgba(240,244,255,0.8)"
_css_uploader_border  = "2px dashed rgba(57,73,171,0.4)" if DM else "2px dashed rgba(57,73,171,0.25)"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

*, .stApp, .block-container, [data-testid="stAppViewContainer"] {{
    font-family: 'Inter', sans-serif !important;
}}

.stApp {{ background: {_bg} !important; }}
.block-container {{ padding-top: 0.6rem; background: {_bg} !important; }}
[data-testid="stSidebar"] {{
    background: {_bg2} !important;
    border-right: 1px solid {_css_sidebar_border} !important;
}}
[data-testid="stSidebar"] * {{ color: {_text} !important; }}

::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: {_bg2}; }}
::-webkit-scrollbar-thumb {{ background: {_accent}; border-radius: 10px; }}
::-webkit-scrollbar-thumb:hover {{ background: #5c6bc0; }}

.dash-header {{
    background: {_css_header_bg};
    padding: 1.2rem 2rem; border-radius: 16px; margin-bottom: 1rem;
    color: white !important; display: flex; align-items: center; gap: 1.5rem;
    box-shadow: {_css_header_shadow};
    animation: slideDown 0.5s cubic-bezier(0.22,1,0.36,1);
    position: relative; overflow: hidden;
}}
.dash-header::before {{
    content: ''; position: absolute; top: -50%; right: -10%;
    width: 400px; height: 400px;
    background: radial-gradient(circle, rgba(121,134,203,0.15) 0%, transparent 70%);
    pointer-events: none;
}}
@keyframes slideDown {{
    from {{ transform: translateY(-20px); opacity: 0; }}
    to   {{ transform: translateY(0);     opacity: 1; }}
}}
/* FIXED: Force white text in header for both modes */
.dash-header h1 {{ margin: 0; font-size: 1.75rem; font-weight: 900; letter-spacing: -0.02em; color: #ffffff !important; text-shadow: 0 1px 4px rgba(0,0,0,0.4); }}
.dash-header p  {{ margin: 0.2rem 0 0 0; opacity: 0.9; font-size: 0.82rem; color: #e8eaf6 !important; }}
.dash-header strong {{ color: #ffffff !important; }}

.kpi-box {{
    background: {_card}; border-radius: 14px; padding: 1.1rem 1.2rem;
    box-shadow: {_css_kpi_shadow};
    text-align: center; border-top: 3px solid {_border};
    transition: transform 0.25s cubic-bezier(0.22,1,0.36,1), box-shadow 0.25s cubic-bezier(0.22,1,0.36,1);
    animation: fadeUp 0.45s cubic-bezier(0.22,1,0.36,1) both;
    cursor: default;
}}
.kpi-box:hover {{
    transform: translateY(-6px) scale(1.01);
    box-shadow: {_css_kpi_hover_shadow};
}}
@keyframes fadeUp {{
    from {{ transform: translateY(16px); opacity: 0; }}
    to   {{ transform: translateY(0);    opacity: 1; }}
}}
.kpi-label {{ font-size: 0.69rem; color: {_sub}; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; }}
.kpi-value {{ font-size: 2rem; font-weight: 900; color: {_text}; line-height: 1.1; letter-spacing: -0.02em; }}
.kpi-sub   {{ font-size: 0.7rem; color: {_sub}; margin-top: 0.25rem; }}

.sec-title {{
    font-size: 0.95rem; font-weight: 700; color: {_text};
    padding: 0.45rem 0; border-bottom: 2px solid {_accent};
    margin-bottom: 0.75rem; letter-spacing: -0.01em;
}}

.profile-card {{
    background: {_card}; border-radius: 16px; padding: 1.3rem 1.6rem;
    border-left: 5px solid {_accent};
    box-shadow: {_css_profile_shadow};
    animation: fadeUp 0.4s cubic-bezier(0.22,1,0.36,1) both;
    transition: box-shadow 0.25s ease;
}}
.profile-card:hover {{
    box-shadow: {_css_profile_hover};
}}
.profile-name {{ font-size: 1.5rem; font-weight: 800; color: {_text}; }}
.profile-sub  {{ font-size: 0.82rem; color: {_sub}; }}

.badge {{
    display: inline-block; padding: 3px 11px; border-radius: 20px;
    font-size: 0.71rem; font-weight: 700; margin: 2px 3px;
    transition: transform 0.15s ease;
}}
.badge:hover {{ transform: scale(1.05); }}
.badge-gold   {{ background: rgba(245,158,11,0.15); color: #f59e0b; border: 1px solid rgba(245,158,11,0.4); }}
.badge-red    {{ background: rgba(239,83,80,0.15);  color: #ef5350; border: 1px solid rgba(239,83,80,0.4); }}
.badge-green  {{ background: rgba(76,175,80,0.15);  color: #4caf50; border: 1px solid rgba(76,175,80,0.4); }}
.badge-blue   {{ background: rgba(121,134,203,0.15); color: {_accent}; border: 1px solid rgba(121,134,203,0.4); }}

/* ── Scrollable tabs ── */
div[data-testid="stTabs"] > div:first-child {{
    overflow-x: auto !important;
    overflow-y: hidden !important;
    flex-wrap: nowrap !important;
    display: flex !important;
    scrollbar-width: thin !important;
    scrollbar-color: {_accent} transparent !important;
    padding-bottom: 2px;
    scroll-behavior: smooth;
    -webkit-overflow-scrolling: touch;
    gap: 2px;
}}
div[data-testid="stTabs"] > div:first-child::-webkit-scrollbar {{
    height: 4px;
}}
div[data-testid="stTabs"] > div:first-child::-webkit-scrollbar-track {{
    background: transparent;
}}
div[data-testid="stTabs"] > div:first-child::-webkit-scrollbar-thumb {{
    background: {_accent};
    border-radius: 10px;
}}
div[data-testid="stTabs"] button {{
    font-weight: 600; font-size: 0.83rem;
    color: {_sub} !important;
    transition: color 0.2s ease, background 0.2s ease;
    border-radius: 6px 6px 0 0;
    white-space: nowrap !important;
    flex-shrink: 0 !important;
}}
div[data-testid="stTabs"] button:hover {{
    color: {_accent} !important;
    background: {_css_tabs_hover_bg} !important;
}}
div[data-testid="stTabs"] button[aria-selected="true"] {{
    color: {_accent} !important;
    border-bottom-color: {_accent} !important;
    background: {_css_tabs_active_bg} !important;
}}

[data-testid="stDataFrame"] {{ border-radius: 12px; overflow: hidden; }}

[data-testid="stTextInput"] input, [data-testid="stSelectbox"] select {{
    background: {_bg2} !important; color: {_text} !important;
    border-color: {_css_input_border} !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
}}
[data-testid="stTextInput"] input:focus {{
    border-color: {_accent} !important;
    box-shadow: 0 0 0 3px rgba(57,73,171,0.2) !important;
}}

p, .stMarkdown, [data-testid="stMarkdownContainer"] * {{ color: {_text}; }}

.rc-banner {{
    background: linear-gradient(135deg, {_css_rc_bg1} 0%, {_css_rc_bg2} 100%);
    border-left: 5px solid #f59e0b; border-radius: 12px;
    padding: 1rem 1.4rem; margin-bottom: 1rem;
    animation: fadeUp 0.4s cubic-bezier(0.22,1,0.36,1) both;
    box-shadow: {_css_rc_shadow};
}}
.rc-issue {{ font-size: 1.05rem; font-weight: 700; color: #f59e0b; }}
.rc-sub   {{ font-size: 0.82rem; color: {_sub}; margin-top: 4px; }}

.top-performer-card {{
    background: {_css_top_bg};
    border: {_css_top_border};
    border-radius: 16px; padding: 1.2rem 1.5rem; margin-bottom: 0.5rem;
    box-shadow: {_css_top_shadow};
    animation: fadeUp 0.4s cubic-bezier(0.22,1,0.36,1) both;
    transition: transform 0.25s ease, box-shadow 0.25s ease;
}}
.top-performer-card:hover {{
    transform: translateY(-4px);
    box-shadow: {_css_top_hover};
}}

.chart-card {{
    background: {_card};
    border-radius: 14px; padding: 0.5rem;
    box-shadow: {_css_chart_shadow};
    transition: box-shadow 0.25s ease, transform 0.25s ease;
    animation: fadeUp 0.5s cubic-bezier(0.22,1,0.36,1) both;
}}
.chart-card:hover {{
    transform: translateY(-3px);
    box-shadow: {_css_chart_hover};
}}

.slow-flag {{
    display: inline-block; padding: 3px 9px;
    background: rgba(239,83,80,0.15); color: #ef5350;
    border: 1px solid rgba(239,83,80,0.4); border-radius: 20px;
    font-size: 0.7rem; font-weight: 700;
}}
.warn-flag {{
    display: inline-block; padding: 3px 9px;
    background: rgba(255,167,38,0.15); color: #ffa726;
    border: 1px solid rgba(255,167,38,0.4); border-radius: 20px;
    font-size: 0.7rem; font-weight: 700;
}}
.ok-flag {{
    display: inline-block; padding: 3px 9px;
    background: rgba(76,175,80,0.15); color: #4caf50;
    border: 1px solid rgba(76,175,80,0.4); border-radius: 20px;
    font-size: 0.7rem; font-weight: 700;
}}

.stButton > button {{
    background: {_css_btn_bg} !important;
    color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 600 !important;
    transition: opacity 0.2s ease, transform 0.15s ease !important;
    box-shadow: 0 2px 12px rgba(57,73,171,0.35) !important;
}}
.stButton > button:hover {{
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 18px rgba(57,73,171,0.5) !important;
}}

[data-testid="stFileUploader"] {{
    background: {_css_uploader_bg} !important;
    border-radius: 12px !important;
    border: {_css_uploader_border} !important;
    transition: border-color 0.2s ease !important;
}}
[data-testid="stFileUploader"]:hover {{
    border-color: {_accent} !important;
}}

.dept-card {{
    border-radius: 14px; padding: 1.1rem 1.3rem; margin-bottom: 0.5rem;
    border-left: 5px solid; transition: transform 0.2s ease, box-shadow 0.2s ease;
    animation: fadeUp 0.4s cubic-bezier(0.22,1,0.36,1) both;
}}
.dept-card:hover {{ transform: translateY(-3px); }}
</style>
""", unsafe_allow_html=True)

# ── Department Station Mappings ─────────────────────────────────────────────
DEPT_STATIONS = {
    "ARSAW": [
        "2319","2320","2327","2328","2335","2336","2341","2342","2348","2349",
        "2356","2357","2362","2363","2368","2369","2376","2377","2382","2383",
        "3317","3318","3325","3326","3333","3334","3339","3340","3346","3347",
        "3355","3356","3361","3362","3367","3368","3376","3377","3382","3383",
        "4320","4321","4328","4329","4335","4336","4341","4342","4348","4349",
        "4356","4357","4363","4364","4369","4370","4377","4378","4385","4386",
    ],
    "PTR": [
        "2118","2119","2125","2126","2134","2135","2141","2142","2147","2148",
        "2152","2153","2159","2160","2165","2166","3118","3119","3125","3126",
        "3134","3135","3141","3142","3147","3148","3153","3154","3159","3160",
        "3166","3167","4118","4119","4125","4126","4137","4138","4144","4145",
        "4150","4151","4155","4156","4162","4163","4169","4170",
    ],
    "ARStow": [
        "2223","2227","2229","2232","2235","2238","2241","2243","2247","2251",
        "2254","2257","2259","2263","2265","2269","3232","3235","3236","3239",
        "3241","3244","3246","3249","3253","3254","3256","3258","3260","3263",
        "3265","3268","4229","4231","4233","4235","4238","4240","4242","4243",
        "4247","4249","4251","4254","4256","4258","4260","4263",
    ],
}
ALL_KNOWN_STATIONS = set()
for _s in DEPT_STATIONS.values():
    ALL_KNOWN_STATIONS.update(_s)

DEPT_COLORS = {
    "ARSAW":     "#7986cb",
    "PTR":       "#66bb6a",
    "ARStow":    "#ffa726",
    "Universal": "#ef5350",
}

def get_department(equipment_id):
    if pd.isna(equipment_id):
        return "Universal"
    eid = str(equipment_id).strip()
    for dept, stations in DEPT_STATIONS.items():
        if eid in stations:
            return dept
    return "Universal"

# ── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"<div style='font-size:1.1rem;font-weight:800;color:{_text};'>⚙️ Settings</div>", unsafe_allow_html=True)
    dm_toggle = st.toggle("🌙 Dark Mode", value=st.session_state.dark_mode, key="dm_toggle_key")
    if dm_toggle != st.session_state.dark_mode:
        st.session_state.dark_mode = dm_toggle
        st.rerun()

    st.markdown("---")
    st.markdown(f"<div style='font-size:0.95rem;font-weight:700;color:{_text};margin-bottom:8px;'>💾 Saved Datasets</div>", unsafe_allow_html=True)
    history_records = history_db.get_history(50)
    if history_records:
        for rec in history_records:
            fhash = rec.get("file_hash", "")
            fname = rec.get("file_name", "unknown")
            display_name = fname[:24] + "…" if len(fname) > 24 else fname
            is_active = fhash in st.session_state.saved_active_hashes
            weeks_str = ", ".join([f"Wk {w}" for w in rec.get("week_numbers", [])]) or "—"
            active_badge = f"<span style='background:rgba(76,175,80,0.2);color:#4caf50;border:1px solid rgba(76,175,80,0.4);border-radius:10px;padding:1px 7px;font-size:0.65rem;font-weight:700;'>ACTIVE</span>" if is_active else ""
            st.markdown(f"""
            <div style="background:{_bg3}; border-radius:10px; padding:8px 10px;
                        margin-bottom:6px; border-left:3px solid {_accent}; font-size:0.74rem;">
                <div style="font-weight:700;color:{_text};white-space:nowrap;overflow:hidden;
                            text-overflow:ellipsis;margin-bottom:3px;" title="{fname}">
                    📄 {display_name} {active_badge}
                </div>
                <div style="color:{_sub};">Saved: {rec.get('upload_ts','')[:16]}</div>
                <div style="color:{_accent};">{rec.get('total_andons',0):,} andons · {weeks_str}</div>
                <div style="color:{_sub};">{rec.get('date_min','')} → {rec.get('date_max','')}</div>
            </div>""", unsafe_allow_html=True)
            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                load_label = "Unload" if is_active else "Load"
                load_icon = "⏏" if is_active else "▶"
                if st.button(f"{load_icon} {load_label}", key=f"load_{fhash}", use_container_width=True):
                    if is_active:
                        st.session_state.saved_active_hashes = [h for h in st.session_state.saved_active_hashes if h != fhash]
                    else:
                        if fhash not in st.session_state.saved_active_hashes:
                            st.session_state.saved_active_hashes.append(fhash)
                    st.rerun()
            with btn_col2:
                if st.button("Remove", key=f"rm_{fhash}", use_container_width=True):
                    history_db.remove_entry(fhash)
                    st.session_state.saved_active_hashes = [h for h in st.session_state.saved_active_hashes if h != fhash]
                    st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Clear All Saved Data", use_container_width=True):
            history_db.clear_history()
            st.session_state.saved_active_hashes = []
            st.rerun()
    else:
        st.markdown(f"<div style='color:{_sub};font-size:0.8rem;padding:6px 0;'>No saved datasets yet.<br>Upload a file to save it.</div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(f"<div style='font-size:0.7rem;color:{_sub};text-align:center;'>LCY3 AFM Dashboard<br>Made by <b>Manish Karki</b></div>", unsafe_allow_html=True)

# FIXED: Header with forced white text always visible in both modes
st.markdown(f"""
<div class="dash-header">
    <div>
        <h1>📊 LCY3 AFM Dashboard — Floor Health</h1>
        <p>Made by <strong>Manish Karki</strong></p>
    </div>
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader("Upload one or more Andon data files (CSV or JSON)", type=["json", "csv"], accept_multiple_files=True)

THRESHOLDS = {
    "Amnesty": 10,
    "Drive Lacking Capability": 10,
}
DEFAULT_THRESHOLD = 5
NO_THRESHOLD = ["Unreachable Charger"]

BLOCKING_EXCLUDE_TYPES = ["Product Problem", "Out of Work"]

def get_threshold(andon_type):
    if andon_type in NO_THRESHOLD:
        return None
    return THRESHOLDS.get(andon_type, DEFAULT_THRESHOLD)


def threshold_color(val, andon_type):
    t = get_threshold(andon_type)
    if pd.isna(val) or t is None:
        return ""
    if val > t * 1.5:
        return "background-color: rgb(210,40,40); color: white; font-weight: 700"
    elif val > t:
        return "background-color: rgb(255,140,0); color: black; font-weight: 700"
    return "background-color: rgb(60,180,60); color: white; font-weight: 700"


CANONICAL_COLS = {
    "status": "Status",
    "resolver": "Resolver",
    "andon type": "Andon Type",
    "dwell time (hh:mm:ss)": "Dwell Time (hh:mm:ss)",
    "time created": "Time Created",
    "equipment type": "Equipment Type",
    "zone": "Zone",
    "shift": "Shift",
    "blocking": "Blocking",
    "equipment id": "Equipment ID",
    "id": "Equipment ID",
    "creator": "Creator",
    "created by": "Creator",
    "time resolved": "Time Resolved",
    "resolved at": "Time Resolved",
}


@st.cache_data
def load_data(file):
    if file.name.endswith(".json"):
        df = pd.read_json(file)
    else:
        df = pd.read_csv(file)
    df.columns = df.columns.str.strip()
    rename_map = {col: CANONICAL_COLS[col.lower()] for col in df.columns if col.lower() in CANONICAL_COLS}
    df = df.rename(columns=rename_map)
    return df


def dwell_color(val, series):
    if pd.isna(val):
        return ""
    valid = series.dropna()
    if len(valid) < 2:
        return ""
    mn, mx = valid.min(), valid.max()
    if mx == mn:
        return ""
    norm = (val - mn) / (mx - mn)
    if norm >= 0.85:
        r, g, b = 210, 40, 40
    elif norm >= 0.65:
        r, g, b = 255, 120, 0
    elif norm >= 0.45:
        r, g, b = 255, 210, 0
    elif norm >= 0.2:
        r, g, b = 130, 220, 130
    else:
        r, g, b = 50, 160, 50
    lum = 0.299 * r + 0.587 * g + 0.114 * b
    fg = "white" if lum < 155 else "black"
    return f"background-color: rgb({r},{g},{b}); color: {fg}; font-weight: 600"


def build_group_pivot(df, group_col):
    cats = sorted(df[group_col].dropna().unique())
    df = df.copy()
    df["DateLabel"] = df["Time Created"].dt.strftime("%b %d, %Y")
    df["Date"] = df["Time Created"].dt.date

    date_count = df.groupby(["Date", "DateLabel", group_col])["Resolve_Min"].count().reset_index()
    date_avg   = df.groupby(["Date", "DateLabel", group_col])["Resolve_Min"].mean().reset_index()

    all_dates = df.groupby(["Date", "DateLabel"]).apply(lambda x: x).reset_index(drop=True)[["Date", "DateLabel"]].drop_duplicates().sort_values("Date", ascending=False)
    date_labels = all_dates.set_index("Date")["DateLabel"].to_dict()

    cols_dict = {}
    for cat in cats:
        c = date_count[date_count[group_col] == cat].set_index("Date")["Resolve_Min"]
        a = date_avg[date_avg[group_col] == cat].set_index("Date")["Resolve_Min"].round(2)
        cols_dict[(cat, "Andons")] = c
        cols_dict[(cat, "Avg Time")] = a

    total_c = df.groupby("Date")["Resolve_Min"].count()
    total_a = df.groupby("Date")["Resolve_Min"].mean().round(2)
    cols_dict[("Total", "Andons")]   = total_c
    cols_dict[("Total", "Avg Time")] = total_a

    tbl = pd.DataFrame(cols_dict).sort_index(ascending=False)
    tbl.index = [date_labels.get(d, str(d)) for d in tbl.index]
    tbl.index.name = "Rows"
    tbl.columns = pd.MultiIndex.from_tuples(tbl.columns)

    grand = {}
    for cat in cats:
        sub = df[df[group_col] == cat]
        grand[(cat, "Andons")]   = int(sub["Resolve_Min"].count())
        grand[(cat, "Avg Time")] = round(sub["Resolve_Min"].mean(), 2)
    grand[("Total", "Andons")]   = int(df["Resolve_Min"].count())
    grand[("Total", "Avg Time")] = round(df["Resolve_Min"].mean(), 2)

    grand_row = pd.DataFrame(grand, index=["Total"])
    grand_row.columns = pd.MultiIndex.from_tuples(grand_row.columns)
    # Grand Total at BOTTOM
    return pd.concat([tbl, grand_row])


def apply_pivot_style(tbl):
    avg_cols = [c for c in tbl.columns if c[1] == "Avg Time"]

    def _style(data):
        s = pd.DataFrame("", index=data.index, columns=data.columns)
        data_rows = [i for i in data.index if i != "Total"]
        for col in avg_cols:
            if col in data.columns:
                series = data.loc[data_rows, col]
                for idx in data_rows:
                    s.loc[idx, col] = dwell_color(data.loc[idx, col], series)
        if "Total" in data.index:
            s.loc["Total"] = "font-weight: 700; background-color: #e8eaf6; color: #1a237e"
        return s

    fmt_avg    = {c: "{:.2f}" for c in tbl.columns if c[1] == "Avg Time"}
    fmt_counts = {c: "{:,.0f}" for c in tbl.columns if c[1] == "Andons"}
    return tbl.style.apply(_style, axis=None).format(fmt_avg, na_rep="—").format(fmt_counts, na_rep="—")


def donut_chart(df, col, title):
    counts = df.groupby(col)["Resolve_Min"].count().reset_index()
    counts.columns = [col, "Count"]
    total = counts["Count"].sum()
    fig = px.pie(counts, names=col, values="Count", hole=0.55,
                 color_discrete_sequence=px.colors.qualitative.Set2)
    fig.update_traces(textinfo="percent+label", pull=[0.03] * len(counts))
    fig.add_annotation(text=f"{total:,}", x=0.5, y=0.5, font_size=22,
                       font_color="#1a237e", showarrow=False,)
    fig.update_layout(showlegend=True, title=title, height=340,
                      legend=dict(orientation="h", yanchor="bottom", y=-0.35),
                      margin=dict(t=40, b=10, l=0, r=0))
    return fig


def hbar_chart(df, col, title):
    avg = (df.groupby(col)["Resolve_Min"].mean()
             .reset_index().sort_values("Resolve_Min")
             .rename(columns={"Resolve_Min": "Avg Time (min)"}))
    fig = px.bar(avg, x="Avg Time (min)", y=col, orientation="h",
                 text=avg["Avg Time (min)"].round(2),
                 color="Avg Time (min)", color_continuous_scale="Blues")
    fig.update_traces(textposition="outside")
    fig.update_layout(coloraxis_showscale=False, title=title, height=340,
                      yaxis_title="", xaxis_title="Avg Dwell Time (min)",
                      margin=dict(t=40, b=10, l=0, r=0))
    return fig


def make_tab_pdf(title, fdf, within_threshold):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"LCY3 AFM Dashboard — {title}", styles["Title"]))
        story.append(Spacer(1, 12))
        total = len(fdf)
        avg_t = fdf["Resolve_Min"].mean()
        med_t = fdf["Resolve_Min"].median()
        within_pct = fdf.apply(within_threshold, axis=1).mean() * 100
        kpi_data = [
            ["Metric", "Value"],
            ["Total Andons", f"{total:,}"],
            ["Avg Resolve Time", f"{avg_t:.2f} min"],
            ["Median Resolve Time", f"{med_t:.2f} min"],
            ["% Within Threshold", f"{within_pct:.1f}%"],
        ]
        t = Table(kpi_data, colWidths=[200, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3949ab")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#f5f5f5"), colors.white]),
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#c5cae9")),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("PADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 16))
        lb = (fdf.groupby("Resolver")
              .agg(Total=("Resolve_Min","count"), Avg=("Resolve_Min","mean"))
              .reset_index().sort_values("Avg").head(20))
        lb["Avg"] = lb["Avg"].round(2)
        lb_data = [["Rank","Resolver","Total Andons","Avg Time (min)"]]
        for i, row in enumerate(lb.itertuples(), 1):
            lb_data.append([str(i), row.Resolver, str(row.Total), f"{row.Avg:.2f}"])
        lt = Table(lb_data, colWidths=[40, 220, 100, 100])
        lt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#f5f5f5"), colors.white]),
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#c5cae9")),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("PADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(Paragraph("Leaderboard (Top 20)", styles["Heading2"]))
        story.append(Spacer(1, 6))
        story.append(lt)
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except ImportError:
        pass

    try:
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, f"LCY3 AFM Dashboard — {title}", ln=True, align="C")
        pdf.ln(4)
        pdf.set_font("Helvetica", size=11)
        total = len(fdf)
        avg_t = fdf["Resolve_Min"].mean()
        med_t = fdf["Resolve_Min"].median()
        within_pct = fdf.apply(within_threshold, axis=1).mean() * 100
        for label, val in [
            ("Total Andons", f"{total:,}"),
            ("Avg Resolve Time", f"{avg_t:.2f} min"),
            ("Median Resolve Time", f"{med_t:.2f} min"),
            ("% Within Threshold", f"{within_pct:.1f}%"),
        ]:
            pdf.cell(90, 8, label, border=1)
            pdf.cell(90, 8, val, border=1, ln=True)
        pdf.ln(6)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Leaderboard (Top 20)", ln=True)
        pdf.set_font("Helvetica", size=9)
        lb = (fdf.groupby("Resolver")
              .agg(Total=("Resolve_Min","count"), Avg=("Resolve_Min","mean"))
              .reset_index().sort_values("Avg").head(20))
        for i, row in enumerate(lb.itertuples(), 1):
            pdf.cell(10, 7, str(i), border=1)
            pdf.cell(90, 7, str(row.Resolver)[:40], border=1)
            pdf.cell(30, 7, str(row.Total), border=1)
            pdf.cell(30, 7, f"{row.Avg:.2f}", border=1, ln=True)
        return pdf.output(dest="S").encode("latin-1")
    except ImportError:
        return None


_parts_new = []
_dup_warnings = []
_new_file_names = []
required_cols = ["Status", "Resolver", "Andon Type", "Dwell Time (hh:mm:ss)", "Time Created"]

if uploaded_files:
    for uf in uploaded_files:
        raw_bytes = uf.read()
        uf.seek(0)
        file_hash = history_db.compute_hash(raw_bytes)

        if history_db.hash_exists(file_hash):
            existing_name = history_db.get_existing_name(file_hash)
            _dup_warnings.append(
                f"**{uf.name}** is identical to a previously saved file "
                f"(**{existing_name}**) — skipped to avoid duplicates. "
                f"Load it from the sidebar instead."
            )
            if file_hash not in st.session_state.saved_active_hashes:
                st.session_state.saved_active_hashes.append(file_hash)
            continue

        part = load_data(uf)
        missing = [c for c in required_cols if c not in part.columns]
        if missing:
            st.error(
                f"**{uf.name}** is missing required column(s): {', '.join(f'`{c}`' for c in missing)}\n\n"
                f"**Columns found:** {', '.join(f'`{c}`' for c in part.columns.tolist())}"
            )
            st.stop()
        part["_source_file"] = uf.name
        _parts_new.append(part)
        _new_file_names.append(uf.name)

        raw_for_preprocess = part.copy()
        raw_for_preprocess["Resolve_Min"] = pd.to_timedelta(
            raw_for_preprocess["Dwell Time (hh:mm:ss)"], errors="coerce"
        ).dt.total_seconds() / 60
        raw_for_preprocess["Time Created"] = pd.to_datetime(
            raw_for_preprocess["Time Created"], errors="coerce"
        )
        raw_for_preprocess = raw_for_preprocess.dropna(subset=["Time Created"])
        raw_for_preprocess["Week"] = raw_for_preprocess["Time Created"].dt.isocalendar().week.astype(int)
        try:
            history_db.record_upload(uf.name, raw_for_preprocess, file_hash)
            if file_hash not in st.session_state.saved_active_hashes:
                st.session_state.saved_active_hashes.append(file_hash)
        except Exception:
            pass

for w in _dup_warnings:
    st.warning(w, icon="⚠️")

_parts_saved = []
for fhash in st.session_state.saved_active_hashes:
    if not any(fhash == history_db.compute_hash(b"") for b in []):
        sdf = history_db.load_dataframe(fhash)
        if sdf is not None and not sdf.empty:
            rec = next((r for r in history_db.get_history() if r.get("file_hash") == fhash), {})
            sdf["_source_file"] = rec.get("file_name", fhash)
            _parts_saved.append(sdf)

_all_parts = _parts_new + _parts_saved

if _all_parts or uploaded_files:
    _usable_parts = [p for p in _all_parts if not p.empty]
    if not _usable_parts:
        st.info("No active data. Load a saved dataset from the sidebar or upload a new file.")
        st.stop()
    df = pd.concat(_usable_parts, ignore_index=True)

    df = df[df["Status"] == "Resolved"].copy()
    df = df[~df["Andon Type"].isin(["Product Problem", "Out of Work"])]
    df["Resolver"] = df["Resolver"].fillna("System").replace("", "System")
    df["Resolve_Min"] = pd.to_timedelta(df["Dwell Time (hh:mm:ss)"], errors="coerce").dt.total_seconds() / 60
    df = df[df["Resolve_Min"].notna()]
    df["Time Created"] = pd.to_datetime(df["Time Created"], errors="coerce")
    df = df[df["Time Created"].notna()]
    df["Date"]  = df["Time Created"].dt.date
    df["Hour"]  = df["Time Created"].dt.hour
    df["Week"]  = df["Time Created"].dt.isocalendar().week.astype(int)

    optional_cols = {
        "Equipment Type": "Equipment Type" in df.columns,
        "Zone":           "Zone"           in df.columns,
        "Shift":          "Shift"          in df.columns,
        "Blocking":       "Blocking"       in df.columns,
        "Equipment ID":   "Equipment ID"   in df.columns,
        "Creator":        "Creator"        in df.columns,
        "Time Resolved":  "Time Resolved"  in df.columns,
    }

    active_source_names = df["_source_file"].unique().tolist() if "_source_file" in df.columns else []
    if active_source_names:
        file_summary_rows = []
        for fname in active_source_names:
            fpart = df[df["_source_file"] == fname]
            if fpart.empty:
                continue
            resolved_count = len(fpart)
            min_dt = fpart["Time Created"].min()
            max_dt = fpart["Time Created"].max()
            file_summary_rows.append((fname, resolved_count, min_dt, max_dt))

        if file_summary_rows:
            card_cols = st.columns(len(file_summary_rows))
            for col_obj, (fname, cnt, dt_min, dt_max) in zip(card_cols, file_summary_rows):
                date_label = (dt_min.strftime("%d %b %Y") if dt_min.date() == dt_max.date()
                              else f"{dt_min.strftime('%d %b')} – {dt_max.strftime('%d %b %Y')}")
                time_label = f"{dt_min.strftime('%H:%M')} → {dt_max.strftime('%H:%M')}"
                display_name = fname if len(fname) <= 28 else fname[:25] + "…"
                col_obj.markdown(f"""
                <div style="background:{_card}; border:1px solid {"rgba(57,73,171,0.25)" if DM else "#c5cae9"};
                            border-left:4px solid {_accent};
                            border-radius:12px; padding:12px 16px; margin-bottom:8px;
                            box-shadow: {"0 2px 16px rgba(0,0,0,0.35)" if DM else "0 2px 10px rgba(0,0,0,0.07)"};
                            transition: transform 0.2s ease, box-shadow 0.2s ease;">
                    <div style="font-size:0.75rem; color:{_accent}; font-weight:700;
                                white-space:nowrap; overflow:hidden; text-overflow:ellipsis;"
                         title="{fname}">📄 {display_name}</div>
                    <div style="font-size:1.5rem; font-weight:900; color:{_text}; line-height:1.2; letter-spacing:-0.02em;">{cnt:,}</div>
                    <div style="font-size:0.72rem; color:{_sub};">andons resolved</div>
                    <div style="font-size:0.72rem; color:{_accent}; margin-top:5px;">📅 {date_label}</div>
                    <div style="font-size:0.72rem; color:{_sub};">⏰ {time_label}</div>
                </div>""", unsafe_allow_html=True)

    f1, f2, f3, f4 = st.columns([2.5, 1.2, 1.2, 1.2])
    with f1:
        min_d, max_d = df["Date"].min(), df["Date"].max()
        date_range = st.date_input("Creation Date", value=(min_d, max_d),
                                   min_value=min_d, max_value=max_d)
    with f2:
        shift_opts = (["All"] + sorted(df["Shift"].dropna().unique().tolist())
                      if optional_cols["Shift"] else ["All"])
        sel_shift = st.selectbox("SHIFT equals", shift_opts)
    with f3:
        resolver_opts = ["All"] + sorted(df["Resolver"].unique().tolist())
        sel_resolver = st.selectbox("Resolver", resolver_opts)
    with f4:
        andon_opts = ["All"] + sorted(df["Andon Type"].dropna().unique().tolist())
        sel_andon = st.selectbox("Andon Type", andon_opts)

    with st.expander("More filters"):
        all_resolvers_list = sorted(df["Resolver"].unique().tolist())
        excluded_resolvers = st.multiselect(
            "Hide resolvers from dashboard",
            options=all_resolvers_list,
            default=[],
            help="Select any logins you want to temporarily hide from all charts and tables."
        )
        search_resolver = st.text_input("🔎 Search resolver by login", "")

    fdf = df.copy()
    if len(date_range) == 2:
        fdf = fdf[(fdf["Date"] >= date_range[0]) & (fdf["Date"] <= date_range[1])]
    if sel_shift != "All" and optional_cols["Shift"]:
        fdf = fdf[fdf["Shift"] == sel_shift]
    if sel_resolver != "All":
        fdf = fdf[fdf["Resolver"] == sel_resolver]
    if sel_andon != "All":
        fdf = fdf[fdf["Andon Type"] == sel_andon]
    if excluded_resolvers:
        fdf = fdf[~fdf["Resolver"].isin(excluded_resolvers)]
    if search_resolver.strip():
        fdf = fdf[fdf["Resolver"].str.contains(search_resolver.strip(), case=False, na=False)]

    st.markdown("<br>", unsafe_allow_html=True)
    k1, k2, k3, k4, k5 = st.columns(5)

    def within_threshold(row):
        t = get_threshold(row["Andon Type"])
        if t is None:
            return True
        return row["Resolve_Min"] <= t

    within_pct = fdf.apply(within_threshold, axis=1).mean() * 100
    efficiency_avg = (fdf.groupby("Resolver").agg(n=("Resolve_Min","count"), avg=("Resolve_Min","mean"))
                     .apply(lambda r: r["n"]/r["avg"] if r["avg"] > 0 else 0, axis=1).mean())

    for col_obj, label, val, sub in [
        (k1, "Total Andons",         f"{len(fdf):,}",                          "Resolved records"),
        (k2, "Avg Resolve Time",     f"{fdf['Resolve_Min'].mean():.2f} min",    "Per andon"),
        (k3, "Median Resolve Time",  f"{fdf['Resolve_Min'].median():.2f} min",  "50th percentile"),
        (k4, "% Within Threshold",   f"{within_pct:.1f}%",                      "vs type targets"),
        (k5, "Avg Efficiency Score", f"{efficiency_avg:.1f}",                   "Andons ÷ Avg Time"),
    ]:
        col_obj.markdown(f"""
        <div class="kpi-box">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{val}</div>
            <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    tab_names = ["📊 Overview", "🏆 Leaderboard", "👤 AFM Profile", "🔍 Root Cause", "AFM Performance", "📋 AFM General", "By Andon Type", "Weekly Breakdown"]
    if optional_cols["Equipment Type"]: tab_names.append("By Equipment Type")
    if optional_cols["Zone"]:           tab_names.append("By Zone")
    if optional_cols["Shift"]:          tab_names.append("By Shift")
    if optional_cols["Blocking"]:       tab_names.append("Blocking Analysis")
    if optional_cols["Blocking"] and optional_cols["Equipment ID"]:
        tab_names.append("🏭 Dept Blocking Analysis")
    if optional_cols["Equipment ID"]:   tab_names.append("Equipment ID Analysis")
    tab_names += ["Hourly Trend", "Heatmap", "Raw Data", "📤 Export", "📂 History"]

    tabs = st.tabs(tab_names)
    tab = {n: t for n, t in zip(tab_names, tabs)}

    def tab_pdf_download(tab_label, df_for_pdf):
        pdf_bytes = make_tab_pdf(tab_label, df_for_pdf, within_threshold)
        if pdf_bytes:
            st.download_button(
                label="⬇️ Download this view as PDF",
                data=pdf_bytes,
                file_name=f"LCY3_{tab_label.replace(' ','_')}.pdf",
                mime="application/pdf",
                key=f"pdf_{tab_label}",
            )
        else:
            st.info("💡 Install `reportlab` or `fpdf2` on your Streamlit Cloud to enable PDF downloads.")

    # ── Column Filter Helper ──────────────────────────────────────────────────
    def col_filters(df_in, tab_key, cols_override=None):
        """Render compact per-column filter widgets; return filtered DataFrame."""
        PRIORITY = ["Andon Type", "Resolver", "Zone", "Shift", "Blocking",
                    "Equipment Type", "Equipment ID", "Creator", "Department", "Week"]
        if cols_override:
            filter_cols = [c for c in cols_override if c in df_in.columns]
        else:
            filter_cols = []
            for col in PRIORITY:
                if col in df_in.columns and 1 < df_in[col].nunique() <= 120:
                    filter_cols.append(col)
        if not filter_cols and "Date" not in df_in.columns:
            return df_in
        with st.expander(f"🔽 Column Filters ({len(filter_cols) + (1 if 'Date' in df_in.columns else 0)} available — click to expand)", expanded=False):
            st.markdown(
                f"<div style='font-size:0.75rem;color:{_sub};margin-bottom:6px;'>"
                "Filter rows in this tab. Leave at <b>All</b> to include everything.</div>",
                unsafe_allow_html=True
            )
            result = df_in.copy()
            if "Date" in df_in.columns:
                d_min, d_max = df_in["Date"].min(), df_in["Date"].max()
                dc1, dc2 = st.columns(2)
                with dc1:
                    tab_date = st.date_input(
                        "📅 Date range",
                        value=(d_min, d_max),
                        min_value=d_min, max_value=d_max,
                        key=f"cf_date_{tab_key}"
                    )
                if len(tab_date) == 2:
                    result = result[(result["Date"] >= tab_date[0]) & (result["Date"] <= tab_date[1])]
            chunk_size = 4
            chunks = [filter_cols[i:i+chunk_size] for i in range(0, len(filter_cols), chunk_size)]
            for chunk in chunks:
                fcols_ui = st.columns(len(chunk))
                for fc_obj, col in zip(fcols_ui, chunk):
                    with fc_obj:
                        opts = ["All"] + sorted(result[col].dropna().astype(str).unique().tolist())
                        chosen = st.selectbox(f"🔎 {col}", options=opts, key=f"cf_{tab_key}_{col}")
                        if chosen != "All":
                            result = result[result[col].astype(str) == chosen]
            removed = len(df_in) - len(result)
            if removed > 0:
                st.markdown(
                    f"<div style='font-size:0.75rem;color:#ffa726;margin-top:4px;'>"
                    f"⚠️ Filters active — showing <b>{len(result):,}</b> of <b>{len(df_in):,}</b> rows "
                    f"({removed:,} hidden)</div>", unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"<div style='font-size:0.75rem;color:#4caf50;margin-top:4px;'>"
                    f"✅ No filters active — showing all <b>{len(result):,}</b> rows</div>",
                    unsafe_allow_html=True
                )
        return result

    # ── Tab: Overview ─────────────────────────────────────────────────────────
    with tab["📊 Overview"]:
        tab_pdf_download("Overview", fdf)
        fdf_ov = col_filters(fdf, "overview")
        st.markdown('<div class="sec-title">Dashboard Overview — Key Insights at a Glance</div>', unsafe_allow_html=True)

        ov1, ov2 = st.columns(2)

        with ov1:
            st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:0.4rem;'>📊 Andons by Resolver (Top 15)</div>", unsafe_allow_html=True)
            top_resolvers = (fdf_ov.groupby("Resolver")["Resolve_Min"].count()
                             .nlargest(15).reset_index()
                             .rename(columns={"Resolve_Min": "Andons"})
                             .sort_values("Andons"))
            colors_bar = []
            max_v = top_resolvers["Andons"].max()
            for v in top_resolvers["Andons"]:
                ratio = v / max_v if max_v > 0 else 0
                if ratio >= 0.85:
                    colors_bar.append("#7986cb")
                elif ratio >= 0.5:
                    colors_bar.append("#5c6bc0")
                else:
                    colors_bar.append("#3949ab")
            fig_ov_bar = go.Figure(go.Bar(
                y=top_resolvers["Resolver"], x=top_resolvers["Andons"],
                orientation="h", marker_color=colors_bar,
                text=top_resolvers["Andons"], textposition="outside",
                hovertemplate="<b>%{y}</b><br>Andons: %{x:,}<extra></extra>"
            ))
            fig_ov_bar.update_layout(
                height=380, xaxis_title="Andon Count", yaxis_title="",
                margin=dict(t=10, b=20, l=0, r=40),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color=_text),
                xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
                yaxis=dict(gridcolor="rgba(0,0,0,0)", color=_text),
            )
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.plotly_chart(fig_ov_bar, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        with ov2:
            st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:0.4rem;'>🥧 Andon Types Distribution</div>", unsafe_allow_html=True)
            type_counts_ov = fdf_ov["Andon Type"].value_counts().reset_index()
            type_counts_ov.columns = ["Andon Type", "Count"]
            fig_ov_pie = px.pie(
                type_counts_ov.head(10), names="Andon Type", values="Count", hole=0.55,
                color_discrete_sequence=["#7986cb","#5c6bc0","#3949ab","#42a5f5","#26c6da",
                                          "#66bb6a","#ffa726","#ef5350","#ab47bc","#26a69a"]
            )
            fig_ov_pie.update_traces(
                textinfo="percent+label", textfont_size=9,
                pull=[0.04 if i == 0 else 0 for i in range(len(type_counts_ov.head(10)))],
                hovertemplate="<b>%{label}</b><br>Count: %{value:,}<br>%{percent}<extra></extra>"
            )
            total_andons_count = int(type_counts_ov["Count"].sum())
            fig_ov_pie.add_annotation(
                text=f"<b>{total_andons_count:,}</b>", x=0.5, y=0.5,
                font_size=20, font_color=_text, showarrow=False
            )
            fig_ov_pie.update_layout(
                height=380, showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.35, font=dict(size=9, color=_text)),
                margin=dict(t=10, b=10, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color=_text)
            )
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.plotly_chart(fig_ov_pie, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin:1rem 0 0.4rem;'>📈 Daily Andon Trend</div>", unsafe_allow_html=True)
        daily_ov = (fdf_ov.groupby("Date").agg(
            Count=("Resolve_Min", "count"),
            Avg=("Resolve_Min", "mean")
        ).reset_index().sort_values("Date"))
        fig_ov_line = go.Figure()
        fig_ov_line.add_trace(go.Scatter(
            x=daily_ov["Date"], y=daily_ov["Count"],
            mode="lines+markers",
            name="Daily Andons",
            line=dict(color="#7986cb", width=2.5, shape="spline", smoothing=0.8),
            marker=dict(size=7, color="#7986cb", line=dict(width=2, color=_bg)),
            fill="tozeroy",
            fillcolor="rgba(121,134,203,0.15)" if DM else "rgba(57,73,171,0.08)",
            hovertemplate="<b>%{x}</b><br>Andons: %{y:,}<extra></extra>"
        ))
        fig_ov_line.add_trace(go.Scatter(
            x=daily_ov["Date"], y=daily_ov["Avg"].round(2),
            mode="lines+markers",
            name="Avg Resolve Time (min)",
            yaxis="y2",
            line=dict(color="#ffa726", width=2, dash="dot"),
            marker=dict(size=5, color="#ffa726"),
            hovertemplate="<b>%{x}</b><br>Avg Time: %{y:.2f} min<extra></extra>"
        ))
        fig_ov_line.add_hline(
            y=DEFAULT_THRESHOLD, line_dash="dash", line_color="#ef5350",
            annotation_text=f"Target ({DEFAULT_THRESHOLD} min)", yref="y2",
            annotation_font_color="#ef5350", annotation_font_size=10
        )
        fig_ov_line.update_layout(
            height=320, xaxis_title="", yaxis_title="Andon Count",
            yaxis2=dict(title="Avg Time (min)", overlaying="y", side="right",
                        showgrid=False, color="#ffa726"),
            margin=dict(t=20, b=40, l=0, r=60),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=10, color=_text)),
            font=dict(color=_text),
            xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
            yaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
        )
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.plotly_chart(fig_ov_line, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if optional_cols["Zone"]:
            st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin:1rem 0 0.4rem;'>🏢 Andons per Zone (Floor)</div>", unsafe_allow_html=True)
            zone_counts = (fdf_ov.groupby("Zone")["Resolve_Min"].count()
                           .reset_index().rename(columns={"Resolve_Min": "Andons"})
                           .sort_values("Andons", ascending=False))
            fig_zone = px.bar(
                zone_counts, x="Zone", y="Andons",
                color="Andons", color_continuous_scale=["#3949ab","#7986cb","#b3baf5"],
                text=zone_counts["Andons"]
            )
            fig_zone.update_traces(textposition="outside",
                                   hovertemplate="<b>%{x}</b><br>Andons: %{y:,}<extra></extra>")
            fig_zone.update_layout(
                height=320, xaxis_title="Zone / Floor", yaxis_title="Andon Count",
                coloraxis_showscale=False,
                margin=dict(t=10, b=40, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color=_text),
                xaxis=dict(gridcolor="rgba(0,0,0,0)", color=_text),
                yaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
            )
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.plotly_chart(fig_zone, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab: Leaderboard ──────────────────────────────────────────────────────
    with tab["🏆 Leaderboard"]:
        tab_pdf_download("Leaderboard", fdf)
        fdf_lb = col_filters(fdf, "leaderboard", cols_override=["Andon Type","Zone","Shift","Date"])
        lb = (fdf_lb.groupby("Resolver")
              .agg(Total_Andons=("Resolve_Min", "count"), Avg_Time=("Resolve_Min", "mean"))
              .reset_index())
        lb["Avg_Time"] = lb["Avg_Time"].round(2)
        lb["Efficiency"] = (lb["Total_Andons"] / lb["Avg_Time"]).round(2)
        lb["Within_Threshold"] = fdf_lb.groupby("Resolver").apply(
            lambda g: g.apply(within_threshold, axis=1).mean() * 100
        ).round(1).values
        lb = lb.sort_values("Avg_Time").reset_index(drop=True)
        lb.index += 1
        lb.index.name = "Rank"

        fastest      = lb.iloc[0]["Resolver"]
        most_active  = lb.nlargest(1, "Total_Andons").iloc[0]["Resolver"]
        most_eff     = lb.nlargest(1, "Efficiency").iloc[0]["Resolver"]

        def assign_badge(r):
            badges = []
            if r == fastest:     badges.append("⚡ Fastest")
            if r == most_active: badges.append("🔥 Most Active")
            if r == most_eff:    badges.append("🎯 Most Efficient")
            return "  ".join(badges) if badges else ""

        lb["Badge"] = lb["Resolver"].apply(assign_badge)

        def flag_slow(row):
            t = get_threshold(None)
            if row["Avg_Time"] > (t or DEFAULT_THRESHOLD) * 1.5:
                return "⚠️ Average"
            elif row["Avg_Time"] > (t or DEFAULT_THRESHOLD):
                return "⚠️ Above target"
            return "✅ On target"

        lb["Status"] = lb.apply(flag_slow, axis=1)

        b1, b2, b3 = st.columns(3)
        for box, icon, title, name, accent_color in [
            (b1, "⚡", "Fastest Resolver",  fastest,     "#f59e0b"),
            (b2, "🔥", "Most Active",       most_active, "#ef5350"),
            (b3, "🎯", "Most Efficient",    most_eff,    "#4caf50"),
        ]:
            stats = lb[lb["Resolver"] == name].iloc[0]
            box.markdown(f"""
            <div class="top-performer-card" style="border-color: rgba({','.join(str(int(accent_color.lstrip('#')[i:i+2],16)) for i in (0,2,4))}, 0.4); border-left: 5px solid {accent_color};">
                <div style="font-size:0.7rem;font-weight:800;text-transform:uppercase;letter-spacing:0.1em;color:{accent_color};margin-bottom:4px;">{icon} {title}</div>
                <div style="font-size:1.35rem;font-weight:900;color:{_text};line-height:1.15;letter-spacing:-0.02em;">⭐ {name}</div>
                <div style="font-size:0.75rem;color:{_sub};margin-top:6px;">
                    <b style="color:{_text};">{stats['Total_Andons']:,}</b> andons &nbsp;·&nbsp;
                    <b style="color:{accent_color};">{stats['Avg_Time']:.2f} min</b> avg
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="sec-title">Full Rankings</div>', unsafe_allow_html=True)

        def style_lb(data):
            s = pd.DataFrame("", index=data.index, columns=data.columns)
            for idx in data.index:
                status = data.loc[idx, "Status"]
                if "Average" in str(status):
                    s.loc[idx, "Avg Time (min)"] = "background-color: rgb(210,40,40); color:white; font-weight:700"
                elif "Above target" in str(status):
                    s.loc[idx, "Avg Time (min)"] = "background-color: rgb(255,140,0); color:black; font-weight:700"
                else:
                    s.loc[idx, "Avg Time (min)"] = "background-color: rgb(60,180,60); color:white; font-weight:700"
                if data.loc[idx, "Badge"]:
                    s.loc[idx, "Badge"] = "background-color: #fef3c7; font-weight:700"
            return s

        display_lb = lb[["Resolver", "Total_Andons", "Avg_Time", "Efficiency", "Within_Threshold", "Badge", "Status"]]
        display_lb.columns = ["Resolver", "Total Andons", "Avg Time (min)", "Efficiency Score", "% Within Threshold", "Badge", "Status"]
        st.dataframe(display_lb.style.apply(style_lb, axis=None)
                     .format({"Avg Time (min)": "{:.2f}", "Efficiency Score": "{:.2f}", "% Within Threshold": "{:.1f}%"}),
                     use_container_width=True, height=450)

        st.markdown('<div class="sec-title">Avg Resolve Time vs Target</div>', unsafe_allow_html=True)
        fig_lb = go.Figure()
        fig_lb.add_trace(go.Bar(
            x=lb["Resolver"], y=lb["Avg_Time"],
            marker_color=["rgb(210,40,40)" if "Average" in s else "rgb(255,140,0)" if "Above target" in s else "rgb(60,180,60)"
                          for s in lb["Status"]],
            text=lb["Avg_Time"].round(2), textposition="outside", name="Avg Time"
        ))
        fig_lb.add_hline(y=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                         annotation_text=f"Default target ({DEFAULT_THRESHOLD} min)")
        fig_lb.update_layout(height=400, xaxis_title="", yaxis_title="Avg Dwell Time (min)",
                             margin=dict(t=30, b=40, l=0, r=0))
        st.plotly_chart(fig_lb, use_container_width=True)

    # ── Tab: AFM Profile ──────────────────────────────────────────────────────
    with tab["👤 AFM Profile"]:
        tab_pdf_download("AFM_Profile", fdf)
        fdf_pr = col_filters(fdf, "profile", cols_override=["Andon Type","Zone","Shift","Date"])
        st.markdown('<div class="sec-title">AFM Resolver Profile — Drill Down</div>', unsafe_allow_html=True)

        resolver_list = sorted(fdf_pr["Resolver"].unique().tolist())
        sel_profile = st.selectbox("Select Resolver", resolver_list, key="profile_sel")
        pdf = fdf_pr[fdf_pr["Resolver"] == sel_profile]

        p_total   = len(pdf)
        p_avg     = pdf["Resolve_Min"].mean()
        p_med     = pdf["Resolve_Min"].median()
        p_within  = pdf.apply(within_threshold, axis=1).mean() * 100
        p_eff     = p_total / p_avg if p_avg > 0 else 0

        daily_p  = pdf.groupby("Date")["Resolve_Min"].agg(Count="count", Avg="mean").reset_index()
        best_day  = daily_p.nsmallest(1, "Avg").iloc[0] if not daily_p.empty else None
        worst_day = daily_p.nlargest(1, "Avg").iloc[0] if not daily_p.empty else None

        is_fastest     = sel_profile == fdf_pr.groupby("Resolver")["Resolve_Min"].mean().idxmin()
        is_most_active = sel_profile == fdf_pr.groupby("Resolver")["Resolve_Min"].count().idxmax()
        badge_html = ""
        if is_fastest:     badge_html += '<span class="badge badge-gold">⚡ Fastest</span>'
        if is_most_active: badge_html += '<span class="badge badge-blue">🔥 Most Active</span>'
        if p_within >= 80: badge_html += '<span class="badge badge-green">✅ On Target</span>'
        if p_avg > DEFAULT_THRESHOLD * 1.5: badge_html += '<span class="badge badge-red">⚠️ Average</span>'

        st.markdown(f"""
        <div class="profile-card">
            <div class="profile-name">🧑‍💻 {sel_profile}</div>
            <div class="profile-sub">Resolver · {p_total:,} andons handled</div>
            <div style="margin-top:8px;">{badge_html}</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        pk1, pk2, pk3, pk4, pk5 = st.columns(5)
        for c_obj, lbl, val, sub in [
            (pk1, "Total Andons",     f"{p_total:,}",         "handled"),
            (pk2, "Avg Time",         f"{p_avg:.2f} min",     "per andon"),
            (pk3, "Median Time",      f"{p_med:.2f} min",     "50th pct"),
            (pk4, "% Within Target",  f"{p_within:.1f}%",     "threshold"),
            (pk5, "Efficiency Score", f"{p_eff:.1f}",         "andons÷avg"),
        ]:
            c_obj.markdown(f"""
            <div class="kpi-box">
                <div class="kpi-label">{lbl}</div>
                <div class="kpi-value">{val}</div>
                <div class="kpi-sub">{sub}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        pc1, pc2 = st.columns(2)

        with pc1:
            st.markdown('<div class="sec-title">📈 Daily Andon Count Trend</div>', unsafe_allow_html=True)
            fig_pt = go.Figure()
            fig_pt.add_trace(go.Scatter(
                x=[str(d) for d in daily_p["Date"]], y=daily_p["Count"],
                mode="lines+markers+text",
                text=daily_p["Count"], textposition="top center",
                line=dict(color=_accent, width=2.5),
                marker=dict(size=8, color=_accent),
                fill="tozeroy",
                fillcolor=f"rgba(57,73,171,{'0.2' if not DM else '0.15'})"
            ))
            fig_pt.update_layout(
                height=320, xaxis_title="", yaxis_title="Andons",
                margin=dict(t=20, b=40, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=False), yaxis=dict(gridcolor="#333" if DM else "#eee")
            )
            st.plotly_chart(fig_pt, use_container_width=True)

        with pc2:
            st.markdown('<div class="sec-title">⏱️ Avg Resolve Time per Day</div>', unsafe_allow_html=True)
            colors_day = ["#ef5350" if v > DEFAULT_THRESHOLD * 1.5
                          else "#ffa726" if v > DEFAULT_THRESHOLD
                          else "#66bb6a" for v in daily_p["Avg"]]
            fig_pa = go.Figure(go.Bar(
                x=[str(d) for d in daily_p["Date"]],
                y=daily_p["Avg"].round(2),
                marker_color=colors_day,
                text=daily_p["Avg"].round(2), textposition="outside"
            ))
            fig_pa.add_hline(y=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                             annotation_text=f"Target ({DEFAULT_THRESHOLD} min)")
            fig_pa.update_layout(
                height=320, xaxis_title="", yaxis_title="Avg Time (min)",
                margin=dict(t=20, b=40, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=False), yaxis=dict(gridcolor="#333" if DM else "#eee")
            )
            st.plotly_chart(fig_pa, use_container_width=True)

        pc3, pc4 = st.columns(2)
        with pc3:
            st.markdown('<div class="sec-title">🥧 Andon Type Breakdown</div>', unsafe_allow_html=True)
            type_p = pdf.groupby("Andon Type")["Resolve_Min"].count().reset_index()
            type_p.columns = ["Andon Type", "Count"]
            fig_pp = px.pie(type_p, names="Andon Type", values="Count", hole=0.5,
                            color_discrete_sequence=px.colors.qualitative.Set2)
            fig_pp.update_layout(height=280, margin=dict(t=20, b=10, l=0, r=0),
                                 paper_bgcolor="rgba(0,0,0,0)",
                                 legend=dict(font_size=9, orientation="h", y=-0.3))
            st.plotly_chart(fig_pp, use_container_width=True)

        with pc4:
            st.markdown('<div class="sec-title">🏅 Best & Worst Days</div>', unsafe_allow_html=True)
            if best_day is not None:
                st.markdown(f"""
                <div style="background:{'#1b2e1b' if DM else '#e8f5e9'};border-left:4px solid #4caf50;
                            border-radius:8px;padding:10px 14px;margin-bottom:8px;">
                    <div style="font-weight:700;color:#4caf50;">🏆 Best Day</div>
                    <div style="font-size:1.1rem;font-weight:800;color:{_text};">{str(best_day['Date'])}</div>
                    <div style="color:{_sub};">Avg: {best_day['Avg']:.2f} min · {int(best_day['Count'])} andons</div>
                </div>""", unsafe_allow_html=True)
            if worst_day is not None:
                st.markdown(f"""
                <div style="background:{'#2e1b1b' if DM else '#ffebee'};border-left:4px solid #ef5350;
                            border-radius:8px;padding:10px 14px;">
                    <div style="font-weight:700;color:#ef5350;">📉 Worst Day</div>
                    <div style="font-size:1.1rem;font-weight:800;color:{_text};">{str(worst_day['Date'])}</div>
                    <div style="color:{_sub};">Avg: {worst_day['Avg']:.2f} min · {int(worst_day['Count'])} andons</div>
                </div>""", unsafe_allow_html=True)

        st.markdown('<div class="sec-title">Full Record Table</div>', unsafe_allow_html=True)
        show_cols = ["Time Created", "Andon Type", "Resolve_Min", "Status"]
        if optional_cols["Zone"]:          show_cols.append("Zone")
        if optional_cols["Equipment ID"]:  show_cols.append("Equipment ID")
        st.dataframe(pdf[show_cols].sort_values("Time Created", ascending=False).rename(
            columns={"Resolve_Min": "Time (min)"}), use_container_width=True, height=340)

    # ── Tab: Root Cause Analysis ───────────────────────────────────────────────
    with tab["🔍 Root Cause"]:
        tab_pdf_download("Root_Cause", fdf)
        fdf_rc = col_filters(fdf, "rootcause")
        st.markdown('<div class="sec-title">Root Cause Analysis — Recurring Issues</div>', unsafe_allow_html=True)

        type_counts = fdf_rc["Andon Type"].value_counts()
        top_issue   = type_counts.index[0]
        top_pct     = type_counts.iloc[0] / len(fdf_rc) * 100

        st.markdown(f"""
        <div class="rc-banner">
            <div class="rc-issue">🚨 Top Recurring Issue: {top_issue} ({top_pct:.1f}%)</div>
            <div class="rc-sub">
                {int(type_counts.iloc[0]):,} of {len(fdf_rc):,} andons · Avg resolve time:
                {fdf[fdf['Andon Type']==top_issue]['Resolve_Min'].mean():.2f} min
            </div>
        </div>""", unsafe_allow_html=True)

        rc1, rc2 = st.columns(2)

        with rc1:
            st.markdown('<div class="sec-title">Top Andon Types by Volume</div>', unsafe_allow_html=True)
            tc_df = type_counts.reset_index()
            tc_df.columns = ["Andon Type", "Count"]
            tc_df["% of Total"] = (tc_df["Count"] / tc_df["Count"].sum() * 100).round(1)
            tc_df["Avg Time (min)"] = tc_df["Andon Type"].map(
                fdf_rc.groupby("Andon Type")["Resolve_Min"].mean().round(2))
            def _get_status(t):
                threshold = get_threshold(t)
                if threshold is None:
                    return "—"
                avg = fdf_rc[fdf_rc["Andon Type"] == t]["Resolve_Min"].mean()
                if avg > threshold * 1.5:
                    return "🚨 Above target"
                elif avg > threshold:
                    return "⚠️ Borderline"
                return "✅ OK"
            tc_df["Status"] = tc_df["Andon Type"].apply(_get_status)
            st.dataframe(tc_df.style.format({"Count": "{:,}", "% of Total": "{:.1f}%",
                                              "Avg Time (min)": "{:.2f}"}),
                         use_container_width=True, height=320)

        with rc2:
            st.markdown('<div class="sec-title">Andon Type Distribution</div>', unsafe_allow_html=True)
            fig_rc_pie = px.pie(tc_df.head(12), names="Andon Type", values="Count", hole=0.5,
                                color_discrete_sequence=px.colors.qualitative.Set3)
            fig_rc_pie.update_traces(textinfo="percent+label", textfont_size=9)
            fig_rc_pie.update_layout(height=340, margin=dict(t=20, b=10, l=0, r=0),
                                     paper_bgcolor="rgba(0,0,0,0)", showlegend=False)
            st.plotly_chart(fig_rc_pie, use_container_width=True)

        rc3, rc4 = st.columns(2)

        with rc3:
            st.markdown('<div class="sec-title">⏰ Peak Hours by Andon Type (Top 5 Types)</div>', unsafe_allow_html=True)
            top5_types = type_counts.head(5).index.tolist()
            hour_type = (fdf_rc[fdf_rc["Andon Type"].isin(top5_types)]
                         .groupby(["Hour", "Andon Type"])["Resolve_Min"].count().reset_index()
                         .rename(columns={"Resolve_Min": "Count"}))
            fig_rc_heat = px.bar(hour_type, x="Hour", y="Count", color="Andon Type",
                                 barmode="stack",
                                 color_discrete_sequence=px.colors.qualitative.Set2,
                                 labels={"Hour": "Hour of Day", "Count": "Andon Count"})
            fig_rc_heat.update_layout(height=320, margin=dict(t=20, b=40, l=0, r=0),
                                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                      legend=dict(font_size=9, orientation="h", y=-0.35))
            st.plotly_chart(fig_rc_heat, use_container_width=True)

        with rc4:
            st.markdown('<div class="sec-title">📅 Week-over-Week Top Issue Trend</div>', unsafe_allow_html=True)
            wow = (fdf_rc[fdf_rc["Andon Type"].isin(top5_types)]
                   .groupby(["Week", "Andon Type"])["Resolve_Min"].count().reset_index()
                   .rename(columns={"Resolve_Min": "Count"}))
            wow["Week"] = "Wk " + wow["Week"].astype(str)
            fig_wow = px.line(wow, x="Week", y="Count", color="Andon Type",
                              markers=True,
                              color_discrete_sequence=px.colors.qualitative.Set2,
                              labels={"Count": "Andon Count", "Week": ""})
            fig_wow.update_layout(height=320, margin=dict(t=20, b=40, l=0, r=0),
                                  paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                  legend=dict(font_size=9, orientation="h", y=-0.35))
            st.plotly_chart(fig_wow, use_container_width=True)

        if optional_cols["Zone"]:
            st.markdown('<div class="sec-title">📍 Top Issue per Zone</div>', unsafe_allow_html=True)
            zone_top = (fdf_rc.groupby(["Zone", "Andon Type"])["Resolve_Min"].count()
                        .reset_index().rename(columns={"Resolve_Min": "Count"})
                        .sort_values("Count", ascending=False)
                        .groupby("Zone").first().reset_index())
            zone_top.columns = ["Zone", "Top Andon Type", "Count"]
            zone_top["% in Zone"] = zone_top.apply(
                lambda r: f"{r['Count'] / fdf_rc[fdf_rc['Zone']==r['Zone']]['Resolve_Min'].count() * 100:.1f}%", axis=1)
            st.dataframe(zone_top, use_container_width=True)

        st.markdown('<div class="sec-title">🧑‍💻 Slowest Resolvers — Flagged Above Threshold</div>', unsafe_allow_html=True)
        slow_df = (fdf_rc.groupby("Resolver")["Resolve_Min"]
                   .agg(Count="count", Avg="mean").reset_index()
                   .sort_values("Avg", ascending=False))
        slow_df["Status"] = slow_df["Avg"].apply(
            lambda x: "⚠️ Average" if x > DEFAULT_THRESHOLD * 1.5
            else ("⚠️ Above target" if x > DEFAULT_THRESHOLD else "✅ OK"))
        slow_df["Avg"] = slow_df["Avg"].round(2)
        slow_flagged = slow_df[slow_df["Status"] != "✅ OK"]
        if not slow_flagged.empty:
            slow_flagged.columns = ["Resolver", "Total Andons", "Avg Time (min)", "Status"]
            st.dataframe(slow_flagged.style.map(
                lambda v: "color: #ef5350; font-weight:700" if "Average" in str(v)
                else ("color: #ffa726; font-weight:700" if "Above target" in str(v) else ""),
                subset=["Status"]
            ), use_container_width=True)
        else:
            st.success("✅ All resolvers are within target thresholds.")

    # ── Tab: AFM Performance ──────────────────────────────────────────────────
    with tab["AFM Performance"]:
        tab_pdf_download("AFM_Performance", fdf)
        fdf_afmperf = col_filters(fdf, "afmperf", cols_override=["Resolver","Zone","Shift","Date"])
        st.markdown('<div class="sec-title">Count and Average Dwell Time by Resolver × Andon Type</div>', unsafe_allow_html=True)

        all_andon_types_afm = sorted(fdf_afmperf["Andon Type"].dropna().unique().tolist())
        sel_andon_types_afm = st.multiselect(
            "🔽 Filter by Andon Type (leave empty = show all)",
            options=all_andon_types_afm,
            default=[],
            key="afm_perf_andon_filter",
            help="Select one or more Andon Types to narrow the table and charts below."
        )
        afm_fdf = fdf_afmperf[fdf_afmperf["Andon Type"].isin(sel_andon_types_afm)] if sel_andon_types_afm else fdf_afmperf.copy()

        if afm_fdf.empty:
            st.warning("No data matches the selected Andon Type filter.")
        else:
            andon_types   = sorted(afm_fdf["Andon Type"].dropna().unique())
            all_resolvers = sorted(afm_fdf["Resolver"].unique())

            cp = afm_fdf.pivot_table(index="Resolver", columns="Andon Type",
                                 values="Resolve_Min", aggfunc="count", fill_value=0).reindex(all_resolvers, fill_value=0)
            ap = afm_fdf.pivot_table(index="Resolver", columns="Andon Type",
                                 values="Resolve_Min", aggfunc="mean").round(2).reindex(all_resolvers)

            afm_cols = {}
            for cat in andon_types:
                afm_cols[(cat, "Andon Count")]    = cp[cat].astype(int)
                afm_cols[(cat, "Dwell Time Avg")] = ap[cat]
            afm_cols[("Total Andons", "Count")]    = cp[andon_types].sum(axis=1).astype(int)
            afm_cols[("Total Andons", "Avg Time")] = afm_fdf.groupby("Resolver")["Resolve_Min"].mean().reindex(all_resolvers).round(2)

            afm_tbl = pd.DataFrame(afm_cols)
            afm_tbl.columns = pd.MultiIndex.from_tuples(afm_tbl.columns)
            afm_tbl.index.name = "AFM"

            grand = {}
            for cat in andon_types:
                sub = afm_fdf[afm_fdf["Andon Type"] == cat]
                grand[(cat, "Andon Count")]    = int(sub["Resolve_Min"].count())
                grand[(cat, "Dwell Time Avg")] = round(sub["Resolve_Min"].mean(), 2)
            grand[("Total Andons", "Count")]    = int(afm_fdf["Resolve_Min"].count())
            grand[("Total Andons", "Avg Time")] = round(afm_fdf["Resolve_Min"].mean(), 2)

            grand_row = pd.DataFrame(grand, index=["Grand Total"])
            grand_row.columns = pd.MultiIndex.from_tuples(grand_row.columns)
            # Grand Total at BOTTOM
            afm_tbl = pd.concat([afm_tbl, grand_row])

            dwell_cs = [(c, "Dwell Time Avg") for c in andon_types] + [("Total Andons", "Avg Time")]

            def _style_afm(data):
                s = pd.DataFrame("", index=data.index, columns=data.columns)
                rows = [i for i in data.index if i != "Grand Total"]
                for col in dwell_cs:
                    if col in data.columns:
                        ser = data.loc[rows, col]
                        for idx in rows:
                            s.loc[idx, col] = dwell_color(data.loc[idx, col], ser)
                if "Grand Total" in data.index:
                    s.loc["Grand Total"] = "font-weight:700; background-color:#e8eaf6; color:#1a237e"
                return s

            afm_styler = (afm_tbl.style.apply(_style_afm, axis=None)
                .format({c: "{:.2f}" for c in afm_tbl.columns if c[1] in ("Dwell Time Avg", "Avg Time")}, na_rep="—")
                .format({c: "{:.0f}" for c in afm_tbl.columns if c[1] in ("Andon Count", "Count")}, na_rep="—"))
            st.dataframe(afm_styler, use_container_width=True, height=450)

            import io as _io
            from openpyxl import Workbook as _WB
            from openpyxl.styles import PatternFill as _PF, Font as _FN, Alignment as _AL, Border as _BD, Side as _SD
            from openpyxl.utils import get_column_letter as _gcl

            def _build_afm_excel(tbl, dwell_cols):
                wb = _WB()
                ws = wb.active
                ws.title = "AFM Performance"
                HDR  = _PF("solid", fgColor="1A237E")
                HFNT = _FN(color="FFFFFF", bold=True, size=9)
                GRN  = _PF("solid", fgColor="E8EAF6")
                GFNT = _FN(color="1A237E", bold=True, size=9)
                RED  = _PF("solid", fgColor="D32F2F")
                ORG  = _PF("solid", fgColor="F57C00")
                GRE  = _PF("solid", fgColor="388E3C")
                YEL  = _PF("solid", fgColor="FFD600")
                LGR  = _PF("solid", fgColor="81C784")
                WFT  = _FN(color="FFFFFF", bold=True, size=9)
                BFT  = _FN(color="000000", bold=True, size=9)
                NFT  = _FN(size=9)
                THIN = _SD(style="thin", color="C5CAE9")
                BDR  = _BD(left=THIN, right=THIN, top=THIN, bottom=THIN)
                CTR  = _AL(horizontal="center", vertical="center")
                c0 = ws.cell(row=1, column=1, value="Resolver")
                c0.fill=HDR; c0.font=HFNT; c0.alignment=CTR; c0.border=BDR
                col_n = 2
                col_map = {}
                cats_seen = {}
                for (cat, sub) in tbl.columns:
                    col_map[(cat, sub)] = col_n
                    if cat not in cats_seen:
                        cats_seen[cat] = col_n
                    col_n += 1
                for cat, start_col in cats_seen.items():
                    sub_count = sum(1 for (c2, _) in tbl.columns if c2 == cat)
                    end_col = start_col + sub_count - 1
                    if start_col != end_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    cell = ws.cell(row=1, column=start_col, value=cat)
                    cell.fill=HDR; cell.font=HFNT; cell.alignment=CTR; cell.border=BDR
                c2h = ws.cell(row=2, column=1, value="Resolver")
                c2h.fill=HDR; c2h.font=HFNT; c2h.alignment=CTR; c2h.border=BDR
                for (cat, sub), cn in col_map.items():
                    cell = ws.cell(row=2, column=cn, value=sub)
                    cell.fill=HDR; cell.font=HFNT; cell.alignment=CTR; cell.border=BDR
                data_rows_idx = [i for i in tbl.index if i != "Grand Total"]
                for r_idx, resolver in enumerate(list(tbl.index), 3):
                    is_grand = resolver == "Grand Total"
                    cell = ws.cell(row=r_idx, column=1, value=resolver)
                    cell.border = BDR
                    if is_grand:
                        cell.fill=GRN; cell.font=GFNT
                    else:
                        cell.font=NFT
                    for (cat, sub), cn in col_map.items():
                        raw = tbl.loc[resolver, (cat, sub)]
                        try:
                            val = float(raw) if not pd.isna(raw) else None
                        except:
                            val = None
                        cell = ws.cell(row=r_idx, column=cn, value=val)
                        cell.border=BDR; cell.alignment=CTR
                        if is_grand:
                            cell.fill=GRN; cell.font=GFNT
                        else:
                            cell.font=NFT
                            if (cat, sub) in dwell_cols and val is not None:
                                sv = [float(tbl.loc[i,(cat,sub)]) for i in data_rows_idx if not pd.isna(tbl.loc[i,(cat,sub)])]
                                if len(sv) >= 2:
                                    mn, mx = min(sv), max(sv)
                                    if mx != mn:
                                        norm = (val-mn)/(mx-mn)
                                        if norm >= 0.85:   cell.fill=RED; cell.font=WFT
                                        elif norm >= 0.65: cell.fill=ORG; cell.font=BFT
                                        elif norm >= 0.45: cell.fill=YEL; cell.font=BFT
                                        elif norm >= 0.2:  cell.fill=LGR; cell.font=BFT
                                        else:              cell.fill=GRE; cell.font=WFT
                for col_obj in ws.columns:
                    mx = 0
                    cl = _gcl(col_obj[0].column)
                    for c in col_obj:
                        try: mx = max(mx, len(str(c.value or "")))
                        except: pass
                    ws.column_dimensions[cl].width = min(mx + 3, 30)
                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            afm_excel = _build_afm_excel(afm_tbl, set(dwell_cs))
            st.download_button(
                "⬇️ Download AFM Performance (.xlsx)",
                afm_excel, "AFM_Performance.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="afm_perf_dl"
            )

            st.markdown("<br>", unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(donut_chart(afm_fdf, "Resolver", "Andons by Resolver"), use_container_width=True)
            with c2:
                st.plotly_chart(hbar_chart(afm_fdf, "Resolver", "Avg Resolve Time by Resolver"), use_container_width=True)

    # ── Tab: AFM General ──────────────────────────────────────────────────────
    with tab["📋 AFM General"]:
        NON_BLOCKING_EXCLUDE = [
            "Replace Fiducial",
            "Untrusted Fiducial Barcode",
            "Out of Work",
            "Product Problem",
            "Unreachable Charger",
            "Drive unit:Repeat offender:Pod barcode failed",
            "Drive Unit:Repeated Offenders: Pod Barcode Failed",
            "Pod Repeated Offender Replace Pod Fiducial",
            "Amnesty",
            "Drive Lacking Capability",
        ]

        THRESH_BLOCKING     = 5
        THRESH_AMNESTY      = 10
        THRESH_DLC          = 10
        THRESH_NON_BLOCKING = 10

        if not optional_cols["Blocking"]:
            st.warning("⚠️ Your data does not have a 'Blocking' column. Please upload a file that includes the Blocking column.")
        else:
            def _norm_bl(v):
                if pd.isna(v):
                    return "no"
                return "yes" if str(v).strip().lower() in ("yes", "true", "1", "y") else "no"

            _base = fdf.copy()
            _base["_BL"] = _base["Blocking"].apply(_norm_bl)
            _base["Resolver"] = _base["Resolver"].fillna("System").replace("", "System")

            st.markdown('<div class="sec-title">📋 AFM General — Performance by Category</div>', unsafe_allow_html=True)

            gf1, gf2, gf3 = st.columns([2, 2, 2])
            with gf1:
                gen_hidden = st.multiselect(
                    "🙈 Hide Resolvers",
                    options=sorted(_base["Resolver"].unique().tolist()),
                    default=[], key="gen_hide_res",
                )
            with gf2:
                gen_show = st.multiselect(
                    "👁️ Show Only These Resolvers",
                    options=sorted(_base["Resolver"].unique().tolist()),
                    default=[], key="gen_show_res",
                )
            with gf3:
                gen_andon_filter = st.multiselect(
                    "🔽 Filter Andon Types",
                    options=sorted(_base["Andon Type"].dropna().unique().tolist()),
                    default=[], key="gen_andon_filter",
                )

            if gen_show:
                _gdf = _base[_base["Resolver"].isin(gen_show)].copy()
            elif gen_hidden:
                _gdf = _base[~_base["Resolver"].isin(gen_hidden)].copy()
            else:
                _gdf = _base.copy()

            if gen_andon_filter:
                _gdf = _gdf[_gdf["Andon Type"].isin(gen_andon_filter)].copy()

            if _gdf.empty:
                st.warning("No data matches the current filters.")
            else:
                all_res = sorted(_gdf["Resolver"].unique())

                am_df  = _gdf[_gdf["Andon Type"] == "Amnesty"].copy()
                dlc_df = _gdf[_gdf["Andon Type"] == "Drive Lacking Capability"].copy()
                bl_df  = _gdf[
                    (_gdf["_BL"] == "yes") &
                    (~_gdf["Andon Type"].isin(["Amnesty", "Drive Lacking Capability"]))
                ].copy()
                nb_df  = _gdf[
                    (_gdf["_BL"] == "no") &
                    (~_gdf["Andon Type"].isin(NON_BLOCKING_EXCLUDE))
                ].copy()

                def _rs(df_s, resolver):
                    r = df_s[df_s["Resolver"] == resolver]
                    c = len(r)
                    a = round(r["Resolve_Min"].mean(), 2) if c > 0 else None
                    return c, a

                rows = {}
                for res in all_res:
                    c_am,  a_am  = _rs(am_df,  res)
                    c_dc,  a_dc  = _rs(dlc_df, res)
                    c_bl,  a_bl  = _rs(bl_df,  res)
                    c_nb,  a_nb  = _rs(nb_df,  res)
                    r_all = _gdf[_gdf["Resolver"] == res]
                    c_tot = len(r_all)
                    a_tot = round(r_all["Resolve_Min"].mean(), 2) if c_tot > 0 else None
                    rows[res] = {
                        ("Amnesty",            "Count"):     c_am,
                        ("Amnesty",            "Avg (min)"): a_am,
                        ("Drive Lacking Cap.", "Count"):     c_dc,
                        ("Drive Lacking Cap.", "Avg (min)"): a_dc,
                        ("Blocking Andons",    "Count"):     c_bl,
                        ("Blocking Andons",    "Avg (min)"): a_bl,
                        ("Non-Blocking",       "Count"):     c_nb,
                        ("Non-Blocking",       "Avg (min)"): a_nb,
                        ("Total Andons",       "Count"):     c_tot,
                        ("Total Andons",       "Avg (min)"): a_tot,
                    }

                gen_tbl = pd.DataFrame(rows).T
                gen_tbl.index.name = "Login"
                gen_tbl.columns = pd.MultiIndex.from_tuples(gen_tbl.columns)

                def _gr(df_s):
                    c = len(df_s)
                    return c, round(df_s["Resolve_Min"].mean(), 2) if c > 0 else None

                c_am_g, a_am_g = _gr(am_df)
                c_dc_g, a_dc_g = _gr(dlc_df)
                c_bl_g, a_bl_g = _gr(bl_df)
                c_nb_g, a_nb_g = _gr(nb_df)
                c_tt_g = len(_gdf)
                a_tt_g = round(_gdf["Resolve_Min"].mean(), 2) if c_tt_g > 0 else None

                grd = pd.DataFrame({
                    ("Amnesty",            "Count"):     [c_am_g],
                    ("Amnesty",            "Avg (min)"): [a_am_g],
                    ("Drive Lacking Cap.", "Count"):     [c_dc_g],
                    ("Drive Lacking Cap.", "Avg (min)"): [a_dc_g],
                    ("Blocking Andons",    "Count"):     [c_bl_g],
                    ("Blocking Andons",    "Avg (min)"): [a_bl_g],
                    ("Non-Blocking",       "Count"):     [c_nb_g],
                    ("Non-Blocking",       "Avg (min)"): [a_nb_g],
                    ("Total Andons",       "Count"):     [c_tt_g],
                    ("Total Andons",       "Avg (min)"): [a_tt_g],
                }, index=["Grand Total"])
                grd.columns = pd.MultiIndex.from_tuples(grd.columns)
                # FIXED: Grand Total at BOTTOM
                gen_tbl = pd.concat([gen_tbl, grd])

                _cat_thresh = {
                    "Amnesty":            THRESH_AMNESTY,
                    "Drive Lacking Cap.": THRESH_DLC,
                    "Blocking Andons":    THRESH_BLOCKING,
                    "Non-Blocking":       THRESH_NON_BLOCKING,
                }

                def _tc(val, cat):
                    t = _cat_thresh.get(cat)
                    if t is None or pd.isna(val):
                        return ""
                    if val >= t:
                        return "background-color:rgb(210,40,40);color:white;font-weight:700"
                    return "background-color:rgb(60,180,60);color:white;font-weight:700"

                avg_g_cols = [c for c in gen_tbl.columns if c[1] == "Avg (min)"]

                def _style_gen(data):
                    s = pd.DataFrame("", index=data.index, columns=data.columns)
                    dr = [i for i in data.index if i != "Grand Total"]
                    for col in avg_g_cols:
                        if col not in data.columns:
                            continue
                        cat = col[0]
                        for idx in dr:
                            val = data.loc[idx, col]
                            try:
                                if pd.isna(val):
                                    continue
                            except Exception:
                                continue
                            s.loc[idx, col] = _tc(val, cat)
                    if "Grand Total" in data.index:
                        s.loc["Grand Total"] = "font-weight:700;background-color:#e8eaf6;color:#1a237e"
                    return s

                gen_styler = (gen_tbl.style
                    .apply(_style_gen, axis=None)
                    .format({c: "{:.2f}" for c in gen_tbl.columns if c[1] == "Avg (min)"}, na_rep="—")
                    .format({c: "{:,.0f}" for c in gen_tbl.columns if c[1] == "Count"}, na_rep="—"))

                st.dataframe(gen_styler, use_container_width=True, height=460)

                st.markdown(f"""
                <div style="display:flex;gap:10px;flex-wrap:wrap;margin:6px 0 12px 0;font-size:0.78rem;">
                    <span style="background:rgb(210,40,40);color:white;padding:2px 10px;border-radius:10px;font-weight:700;">🔴 Blocking avg ≥ {THRESH_BLOCKING} min</span>
                    <span style="background:rgb(210,40,40);color:white;padding:2px 10px;border-radius:10px;font-weight:700;">🔴 Amnesty avg ≥ {THRESH_AMNESTY} min</span>
                    <span style="background:rgb(210,40,40);color:white;padding:2px 10px;border-radius:10px;font-weight:700;">🔴 DLC avg ≥ {THRESH_DLC} min</span>
                    <span style="background:rgb(210,40,40);color:white;padding:2px 10px;border-radius:10px;font-weight:700;">🔴 Non-Blocking avg ≥ {THRESH_NON_BLOCKING} min</span>
                    <span style="background:rgb(60,180,60);color:white;padding:2px 10px;border-radius:10px;font-weight:700;">🟢 Below threshold</span>
                </div>
                """, unsafe_allow_html=True)

                st.markdown('<div class="sec-title">⏱️ Hours Lost to Blocking Andons</div>', unsafe_allow_html=True)
                st.caption("Hours lost = total dwell time of Blocking andons — time the station was stopped waiting for resolution.")

                today_date = _gdf["Date"].max()
                t_bl_today = _gdf[
                    (_gdf["Date"] == today_date) &
                    (_gdf["_BL"] == "yes") &
                    (~_gdf["Andon Type"].isin(["Amnesty", "Drive Lacking Capability"]))
                ]
                t_min  = t_bl_today["Resolve_Min"].sum()
                t_hrs  = t_min / 60
                t_cnt  = len(t_bl_today)
                tot_hrs = bl_df["Resolve_Min"].sum() / 60

                hl1, hl2, hl3, hl4 = st.columns(4)
                for box, lbl, val, sub, color in [
                    (hl1, "Latest Day Hours Lost", f"{t_hrs:.2f} hrs",        f"{t_min:.0f} min | {str(today_date)}", "#ef5350"),
                    (hl2, "Latest Day Blocking",   f"{t_cnt:,}",               "blocking andons",                      "#ffa726"),
                    (hl3, "Total Hours Lost",       f"{tot_hrs:.2f} hrs",      "across all dates",                     "#3949ab"),
                    (hl4, "Total Blocking Andons",  f"{len(bl_df):,}",          "in filtered range",                    "#3949ab"),
                ]:
                    box.markdown(f"""
                    <div class="kpi-box" style="border-top-color:{color};">
                        <div class="kpi-label">{lbl}</div>
                        <div class="kpi-value" style="font-size:1.5rem;color:{color};">{val}</div>
                        <div class="kpi-sub">{sub}</div>
                    </div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                daily_bl_trend = (
                    bl_df.groupby(["Date", "Andon Type"])
                    .agg(Minutes_Lost=("Resolve_Min","sum"), Count=("Resolve_Min","count"))
                    .reset_index()
                )
                daily_bl_trend["Hours_Lost"] = (daily_bl_trend["Minutes_Lost"] / 60).round(3)
                daily_bl_total = (
                    bl_df.groupby("Date")
                    .agg(
                        Hours_Total=("Resolve_Min", lambda x: round(x.sum()/60, 3)),
                        Count_Total=("Resolve_Min", "count")
                    ).reset_index()
                )

                if not daily_bl_trend.empty:
                    hl_c1, hl_c2 = st.columns([2, 1])
                    with hl_c1:
                        st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>Daily Hours Lost by Blocking Andon Type</div>", unsafe_allow_html=True)
                        fig_hl = px.bar(
                            daily_bl_trend, x="Date", y="Hours_Lost", color="Andon Type",
                            barmode="stack",
                            color_discrete_sequence=px.colors.qualitative.Set2,
                            labels={"Hours_Lost": "Hours Lost", "Date": ""}
                        )
                        fig_hl.add_trace(go.Scatter(
                            x=daily_bl_total["Date"].astype(str),
                            y=daily_bl_total["Count_Total"],
                            name="Blocking Count", yaxis="y2",
                            mode="lines+markers",
                            line=dict(color="#1a237e", width=2),
                            marker=dict(size=6)
                        ))
                        fig_hl.update_layout(
                            height=360,
                            yaxis=dict(title="Hours Lost"),
                            yaxis2=dict(title="Blocking Count", overlaying="y", side="right", showgrid=False),
                            legend=dict(orientation="h", y=-0.35, font_size=9),
                            margin=dict(t=10, b=10, l=0, r=0),
                            paper_bgcolor="rgba(0,0,0,0)",
                            plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=_text)
                        )
                        st.plotly_chart(fig_hl, use_container_width=True)
                    with hl_c2:
                        st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>Summary by Type</div>", unsafe_allow_html=True)
                        hl_summary = (
                            bl_df.groupby("Andon Type")
                            .agg(Count=("Resolve_Min","count"), Total_Min=("Resolve_Min","sum"), Avg_Min=("Resolve_Min","mean"))
                            .reset_index()
                            .assign(
                                Hours_Lost=lambda d: (d["Total_Min"]/60).round(2),
                                Total_Min=lambda d: d["Total_Min"].round(1),
                                Avg_Min=lambda d: d["Avg_Min"].round(2)
                            )
                            .sort_values("Hours_Lost", ascending=False)
                            .rename(columns={"Total_Min":"Total Min","Avg_Min":"Avg Min","Hours_Lost":"Hrs Lost"})
                        )
                        st.dataframe(
                            hl_summary[["Andon Type","Count","Hrs Lost","Avg Min"]],
                            use_container_width=True, hide_index=True, height=340
                        )
                else:
                    st.info("No blocking andons found in the current filter range.")

                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown('<div class="sec-title">⬇️ Download AFM General Report</div>', unsafe_allow_html=True)
                dl_g1, dl_g2 = st.columns(2)

                import io as _io2
                from openpyxl import Workbook as _WB2
                from openpyxl.styles import (PatternFill as _PF2, Font as _FN2,
                                              Alignment as _AL2, Border as _BD2, Side as _SD2)
                from openpyxl.utils import get_column_letter as _gcl2

                def _build_gen_excel(tbl, gfdf):
                    wb = _WB2()
                    ws = wb.active
                    ws.title = "AFM General"
                    HDR  = _PF2("solid", fgColor="1A237E")
                    HFNT = _FN2(color="FFFFFF", bold=True, size=9)
                    GRFL = _PF2("solid", fgColor="E8EAF6")
                    GRFN = _FN2(color="1A237E", bold=True, size=9)
                    RED  = _PF2("solid", fgColor="D32F2F")
                    GRE  = _PF2("solid", fgColor="388E3C")
                    WFT  = _FN2(color="FFFFFF", bold=True, size=9)
                    NFT  = _FN2(size=9)
                    WHT  = _PF2("solid", fgColor="FFFFFF")
                    THIN = _SD2(style="thin", color="C5CAE9")
                    BDR  = _BD2(left=THIN, right=THIN, top=THIN, bottom=THIN)
                    CTR  = _AL2(horizontal="center", vertical="center")
                    _exc_thresh = {
                        "Amnesty":            THRESH_AMNESTY,
                        "Drive Lacking Cap.": THRESH_DLC,
                        "Blocking Andons":    THRESH_BLOCKING,
                        "Non-Blocking":       THRESH_NON_BLOCKING,
                    }
                    c0 = ws.cell(row=1, column=1, value="Login")
                    c0.fill = HDR; c0.font = HFNT; c0.alignment = CTR; c0.border = BDR
                    col_map = {}
                    cats_seen = {}
                    cn = 2
                    for (cat, sub) in tbl.columns:
                        col_map[(cat, sub)] = cn
                        if cat not in cats_seen:
                            cats_seen[cat] = cn
                        cn += 1
                    for cat, start_c in cats_seen.items():
                        sub_count = sum(1 for (c2, _) in tbl.columns if c2 == cat)
                        end_c = start_c + sub_count - 1
                        if start_c != end_c:
                            ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
                        cell = ws.cell(row=1, column=start_c, value=cat)
                        cell.fill = HDR; cell.font = HFNT; cell.alignment = CTR; cell.border = BDR
                    ws.cell(row=2, column=1, value="Login").fill = HDR
                    ws.cell(row=2, column=1).font = HFNT
                    ws.cell(row=2, column=1).alignment = CTR
                    ws.cell(row=2, column=1).border = BDR
                    for (cat, sub), c_n in col_map.items():
                        cell = ws.cell(row=2, column=c_n, value=sub)
                        cell.fill = HDR; cell.font = HFNT; cell.alignment = CTR; cell.border = BDR
                    data_rows_idx = [i for i in tbl.index if i != "Grand Total"]
                    for r_idx, resolver in enumerate(tbl.index, 3):
                        is_grand = resolver == "Grand Total"
                        cell = ws.cell(row=r_idx, column=1, value=resolver)
                        cell.border = BDR
                        if is_grand:
                            cell.fill = GRFL; cell.font = GRFN
                        else:
                            cell.fill = WHT; cell.font = NFT
                        for (cat, sub), c_n in col_map.items():
                            raw = tbl.loc[resolver, (cat, sub)]
                            try:
                                val = float(raw) if not pd.isna(raw) else None
                            except Exception:
                                val = None
                            cell = ws.cell(row=r_idx, column=c_n, value=val)
                            cell.border = BDR; cell.alignment = CTR
                            if is_grand:
                                cell.fill = GRFL; cell.font = GRFN
                            elif sub == "Avg (min)" and val is not None and cat in _exc_thresh:
                                t = _exc_thresh[cat]
                                if val >= t:
                                    cell.fill = RED; cell.font = WFT
                                else:
                                    cell.fill = GRE; cell.font = WFT
                            else:
                                cell.fill = WHT; cell.font = NFT
                    for col_obj in ws.columns:
                        mx = 0
                        cl = _gcl2(col_obj[0].column)
                        for c in col_obj:
                            try:
                                mx = max(mx, len(str(c.value or "")))
                            except Exception:
                                pass
                        ws.column_dimensions[cl].width = min(mx + 3, 28)
                    ws2 = wb.create_sheet("Hours Lost")
                    for i, hdr in enumerate(["Date","Blocking Andon Type","Resolver","Count","Minutes Lost","Hours Lost"], 1):
                        c2 = ws2.cell(row=1, column=i, value=hdr)
                        c2.fill = HDR; c2.font = HFNT; c2.border = BDR; c2.alignment = CTR
                    detail_bl = (
                        gfdf[
                            (gfdf["_BL"] == "yes") &
                            (~gfdf["Andon Type"].isin(["Amnesty","Drive Lacking Capability"]))
                        ]
                        .groupby(["Date","Andon Type","Resolver"])
                        .agg(Count=("Resolve_Min","count"), Minutes=("Resolve_Min","sum"))
                        .reset_index()
                    )
                    detail_bl["Hours"]   = (detail_bl["Minutes"]/60).round(3)
                    detail_bl["Minutes"] = detail_bl["Minutes"].round(2)
                    detail_bl = detail_bl.sort_values(["Date","Andon Type","Resolver"])
                    for ri, row in detail_bl.iterrows():
                        ws2.cell(row=ri+2, column=1, value=str(row["Date"]))
                        ws2.cell(row=ri+2, column=2, value=row["Andon Type"])
                        ws2.cell(row=ri+2, column=3, value=row["Resolver"])
                        ws2.cell(row=ri+2, column=4, value=int(row["Count"]))
                        ws2.cell(row=ri+2, column=5, value=row["Minutes"])
                        ws2.cell(row=ri+2, column=6, value=row["Hours"])
                    for col_obj in ws2.columns:
                        mx = 0
                        cl = _gcl2(col_obj[0].column)
                        for c in col_obj:
                            try:
                                mx = max(mx, len(str(c.value or "")))
                            except Exception:
                                pass
                        ws2.column_dimensions[cl].width = min(mx + 4, 30)
                    buf2 = _io2.BytesIO()
                    wb.save(buf2)
                    buf2.seek(0)
                    return buf2.getvalue()

                with dl_g1:
                    gen_excel = _build_gen_excel(gen_tbl, _gdf)
                    st.download_button(
                        "⬇️ Download AFM General (.xlsx)",
                        gen_excel, "AFM_General.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="gen_excel_dl"
                    )
                with dl_g2:
                    try:
                        import pdf_report as _prg
                        gen_pdf = _prg.build_pdf_daily(fdf, uploaded_files, within_threshold)
                        st.download_button(
                            "⬇️ Download PDF Report",
                            gen_pdf, "LCY3_AFM_General.pdf",
                            "application/pdf",
                            use_container_width=True, key="gen_pdf_dl"
                        )
                    except Exception:
                        st.caption("Add pdf_report.py to enable PDF export.")

    # ── Tab: By Andon Type ────────────────────────────────────────────────────
    with tab["By Andon Type"]:
        tab_pdf_download("By_Andon_Type", fdf)
        fdf_at = col_filters(fdf, "bytype", cols_override=["Resolver","Zone","Shift","Date"])
        st.markdown('<div class="sec-title">Number of Andons and Dwell Time by Date × Andon Type</div>', unsafe_allow_html=True)
        tbl_at = build_group_pivot(fdf_at, "Andon Type")
        st.dataframe(apply_pivot_style(tbl_at), use_container_width=True, height=400)
        c1, c2 = st.columns(2)
        with c1:
            st.plotly_chart(donut_chart(fdf_at, "Andon Type", "Andons by Type"), use_container_width=True)
        with c2:
            st.plotly_chart(hbar_chart(fdf_at, "Andon Type", "Avg Resolve Time by Andon Type"), use_container_width=True)

    # ── Tab: Weekly Breakdown ─────────────────────────────────────────────────
    with tab["Weekly Breakdown"]:
        tab_pdf_download("Weekly_Breakdown", fdf)
        fdf_wk = col_filters(fdf, "weekly", cols_override=["Andon Type","Resolver","Zone","Shift"])
        weeks_avail = sorted(fdf_wk["Week"].dropna().unique(), reverse=True)

        st.markdown('<div class="sec-title">Andons by Type and Week</div>', unsafe_allow_html=True)
        week_count_p = fdf_wk.pivot_table(index="Andon Type", columns="Week",
                                       values="Resolve_Min", aggfunc="count", fill_value=0)
        week_avg_p   = fdf_wk.pivot_table(index="Andon Type", columns="Week",
                                       values="Resolve_Min", aggfunc="mean")

        week_cols = {}
        for w in weeks_avail:
            if w in week_count_p.columns:
                week_cols[(f"Wk {w}", "Andons")]   = week_count_p[w].astype(int)
                week_cols[(f"Wk {w}", "Avg Time")] = week_avg_p[w].round(2) if w in week_avg_p.columns else pd.Series(dtype=float)
        week_cols[("Total", "Andons")]   = week_count_p[weeks_avail].sum(axis=1).astype(int)
        week_cols[("Total", "Avg Time")] = fdf.groupby("Andon Type")["Resolve_Min"].mean().round(2)

        weekly_tbl = pd.DataFrame(week_cols)
        weekly_tbl.columns = pd.MultiIndex.from_tuples(weekly_tbl.columns)
        weekly_tbl.index.name = "Andon Type"
        weekly_tbl = weekly_tbl.sort_values(("Total", "Andons"), ascending=False)

        wk_grand = {}
        for w in weeks_avail:
            sub_w = fdf_wk[fdf_wk["Week"] == w]
            wk_grand[(f"Wk {w}", "Andons")]   = int(sub_w["Resolve_Min"].count())
            wk_grand[(f"Wk {w}", "Avg Time")] = round(sub_w["Resolve_Min"].mean(), 2)
        wk_grand[("Total", "Andons")]   = int(fdf_wk["Resolve_Min"].count())
        wk_grand[("Total", "Avg Time")] = round(fdf_wk["Resolve_Min"].mean(), 2)

        wk_grand_row = pd.DataFrame(wk_grand, index=["Grand Total"])
        wk_grand_row.columns = pd.MultiIndex.from_tuples(wk_grand_row.columns)
        # Grand Total at BOTTOM
        weekly_tbl = pd.concat([weekly_tbl, wk_grand_row])

        wk_avg_cols = [c for c in weekly_tbl.columns if c[1] == "Avg Time"]

        def _style_weekly(data):
            s = pd.DataFrame("", index=data.index, columns=data.columns)
            data_rows = [i for i in data.index if i != "Grand Total"]
            for col in wk_avg_cols:
                if col in data.columns:
                    ser = data.loc[data_rows, col]
                    for idx in data_rows:
                        s.loc[idx, col] = dwell_color(data.loc[idx, col], ser)
            if "Grand Total" in data.index:
                s.loc["Grand Total"] = "font-weight:700; background-color:#e8eaf6; color:#1a237e"
            return s

        wk_styler = (weekly_tbl.style.apply(_style_weekly, axis=None)
            .format({c: "{:.2f}" for c in weekly_tbl.columns if c[1] == "Avg Time"}, na_rep="—")
            .format({c: "{:,.0f}" for c in weekly_tbl.columns if c[1] == "Andons"}, na_rep="—"))
        st.dataframe(wk_styler, use_container_width=True, height=420)

        st.markdown('<div class="sec-title">AFM Performance by Week (Resolver × Week)</div>', unsafe_allow_html=True)

        afm_wk_cnt = fdf_wk.pivot_table(index="Resolver", columns="Week",
                                     values="Resolve_Min", aggfunc="count", fill_value=0)
        afm_wk_avg = fdf_wk.pivot_table(index="Resolver", columns="Week",
                                     values="Resolve_Min", aggfunc="mean")

        afm_wk_cols = {}
        for w in weeks_avail:
            if w in afm_wk_cnt.columns:
                afm_wk_cols[(f"Wk {w}", "Andons")]   = afm_wk_cnt[w].astype(int)
                afm_wk_cols[(f"Wk {w}", "Avg Time")] = afm_wk_avg[w].round(2) if w in afm_wk_avg.columns else pd.Series(dtype=float)
        afm_wk_cols[("Total", "Andons")]   = afm_wk_cnt[weeks_avail].sum(axis=1).astype(int)
        afm_wk_cols[("Total", "Avg Time")] = fdf.groupby("Resolver")["Resolve_Min"].mean().round(2)

        afm_wk_tbl = pd.DataFrame(afm_wk_cols)
        afm_wk_tbl.columns = pd.MultiIndex.from_tuples(afm_wk_tbl.columns)
        afm_wk_tbl.index.name = "Resolver"
        afm_wk_tbl = afm_wk_tbl.sort_values(("Total", "Andons"), ascending=False)

        afm_wk_grand = {}
        for w in weeks_avail:
            sub_w = fdf[fdf["Week"] == w]
            afm_wk_grand[(f"Wk {w}", "Andons")]   = int(sub_w["Resolve_Min"].count())
            afm_wk_grand[(f"Wk {w}", "Avg Time")] = round(sub_w["Resolve_Min"].mean(), 2)
        afm_wk_grand[("Total", "Andons")]   = int(fdf["Resolve_Min"].count())
        afm_wk_grand[("Total", "Avg Time")] = round(fdf["Resolve_Min"].mean(), 2)

        afm_wk_grand_row = pd.DataFrame(afm_wk_grand, index=["Grand Total"])
        afm_wk_grand_row.columns = pd.MultiIndex.from_tuples(afm_wk_grand_row.columns)
        # Grand Total at BOTTOM
        afm_wk_tbl = pd.concat([afm_wk_tbl, afm_wk_grand_row])

        afm_wk_avg_cols = [c for c in afm_wk_tbl.columns if c[1] == "Avg Time"]

        def _style_afm_wk(data):
            s = pd.DataFrame("", index=data.index, columns=data.columns)
            data_rows = [i for i in data.index if i != "Grand Total"]
            for col in afm_wk_avg_cols:
                if col in data.columns:
                    ser = data.loc[data_rows, col]
                    for idx in data_rows:
                        s.loc[idx, col] = dwell_color(data.loc[idx, col], ser)
            if "Grand Total" in data.index:
                s.loc["Grand Total"] = "font-weight:700; background-color:#e8eaf6; color:#1a237e"
            return s

        afm_wk_styler = (afm_wk_tbl.style.apply(_style_afm_wk, axis=None)
            .format({c: "{:.2f}" for c in afm_wk_tbl.columns if c[1] == "Avg Time"}, na_rep="—")
            .format({c: "{:,.0f}" for c in afm_wk_tbl.columns if c[1] == "Andons"}, na_rep="—"))
        st.dataframe(afm_wk_styler, use_container_width=True, height=420)

        st.markdown('<div class="sec-title">Avg Resolve Time per Resolver</div>', unsafe_allow_html=True)
        afm_bar = (fdf.groupby("Resolver")["Resolve_Min"]
                   .agg(Avg="mean", Count="count").reset_index()
                   .sort_values("Avg"))
        afm_bar["Color"] = afm_bar["Avg"].apply(
            lambda x: "rgb(60,180,60)" if x <= DEFAULT_THRESHOLD
            else ("rgb(255,140,0)" if x <= DEFAULT_THRESHOLD * 1.5 else "rgb(210,40,40)"))
        fig_afm_bar = go.Figure(go.Bar(
            x=afm_bar["Resolver"], y=afm_bar["Avg"],
            marker_color=afm_bar["Color"],
            text=afm_bar["Avg"].round(2), textposition="outside",
            customdata=afm_bar["Count"],
            hovertemplate="<b>%{x}</b><br>Avg: %{y:.2f} min<br>Andons: %{customdata}<extra></extra>"
        ))
        fig_afm_bar.add_hline(y=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                               annotation_text=f"Default target ({DEFAULT_THRESHOLD} min)")
        fig_afm_bar.update_layout(height=380, xaxis_title="", yaxis_title="Avg Dwell Time (min)",
                                  margin=dict(t=30, b=40, l=0, r=0))
        st.plotly_chart(fig_afm_bar, use_container_width=True)

        st.markdown('<div class="sec-title">System vs Non-System Andons by Week</div>', unsafe_allow_html=True)
        fdf["_resolver_type"] = fdf["Resolver"].apply(lambda r: "System" if r == "System" else "Non-System")
        sys_wk = (fdf.groupby(["Week", "_resolver_type"])["Resolve_Min"]
                  .agg(Andons="count", Avg_Time="mean").reset_index())
        sys_wk["Week_Label"] = "Wk " + sys_wk["Week"].astype(str)

        col_sys, col_nonsys = st.columns(2)
        with col_sys:
            fig_sys_bar = px.bar(sys_wk, x="Week_Label", y="Andons", color="_resolver_type",
                                 barmode="group",
                                 color_discrete_map={"System": "#ef5350", "Non-System": "#42a5f5"},
                                 title="Total Andons: System vs Non-System per Week",
                                 labels={"_resolver_type": "Resolver Type", "Week_Label": "Week", "Andons": "Count"})
            fig_sys_bar.update_layout(height=360, margin=dict(t=50, b=30, l=0, r=0),
                                      legend=dict(orientation="h", y=-0.25))
            st.plotly_chart(fig_sys_bar, use_container_width=True)

        with col_nonsys:
            fig_sys_avg = px.bar(sys_wk, x="Week_Label", y="Avg_Time", color="_resolver_type",
                                 barmode="group",
                                 color_discrete_map={"System": "#ef5350", "Non-System": "#42a5f5"},
                                 title="Avg Resolve Time: System vs Non-System per Week",
                                 labels={"_resolver_type": "Resolver Type", "Week_Label": "Week", "Avg_Time": "Avg Time (min)"})
            fig_sys_avg.add_hline(y=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                                   annotation_text=f"Target ({DEFAULT_THRESHOLD} min)")
            fig_sys_avg.update_layout(height=360, margin=dict(t=50, b=30, l=0, r=0),
                                       legend=dict(orientation="h", y=-0.25))
            st.plotly_chart(fig_sys_avg, use_container_width=True)

        sys_summary = (fdf.groupby("_resolver_type")["Resolve_Min"]
                       .agg(Total_Andons="count", Avg_Time="mean").reset_index()
                       .rename(columns={"_resolver_type": "Resolver Type",
                                        "Total_Andons": "Total Andons", "Avg_Time": "Avg Time (min)"}))
        sys_summary["Avg Time (min)"] = sys_summary["Avg Time (min)"].round(2)
        sys_summary["% of Total"] = (sys_summary["Total Andons"] / sys_summary["Total Andons"].sum() * 100).round(1).astype(str) + "%"
        st.dataframe(sys_summary.set_index("Resolver Type"), use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            top10 = (fdf.groupby("Andon Type")["Resolve_Min"].count()
                     .nlargest(10).reset_index().rename(columns={"Resolve_Min": "Count"}))
            fig_wk_d = px.pie(top10, names="Andon Type", values="Count", hole=0.55,
                              color_discrete_sequence=px.colors.qualitative.Set2,
                              title="Count of Problem Id by Andon Type (Top 10)")
            fig_wk_d.update_traces(textinfo="percent+label")
            fig_wk_d.add_annotation(text=f"{fdf['Resolve_Min'].count():,}", x=0.5, y=0.5,
                                    font_size=22, font_color="#1a237e", showarrow=False)
            fig_wk_d.update_layout(height=380, legend=dict(orientation="h", y=-0.35),
                                   margin=dict(t=50, b=10, l=0, r=0))
            st.plotly_chart(fig_wk_d, use_container_width=True)
        with c2:
            st.plotly_chart(hbar_chart(fdf, "Andon Type", "Average of Dwell Time Minutes by Andon Type"),
                            use_container_width=True)

    # ── Optional tabs ─────────────────────────────────────────────────────────
    if optional_cols["Equipment Type"]:
        with tab["By Equipment Type"]:
            tab_pdf_download("By_Equipment_Type", fdf)
            fdf_et = col_filters(fdf, "eqtype", cols_override=["Resolver","Zone","Shift","Andon Type","Date"])
            st.markdown('<div class="sec-title">Number of Andons and Resolution Times by Equipment Type</div>', unsafe_allow_html=True)
            tbl_et = build_group_pivot(fdf_et, "Equipment Type")
            st.dataframe(apply_pivot_style(tbl_et), use_container_width=True, height=400)
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(donut_chart(fdf_et, "Equipment Type", "Andons by Equipment Type"), use_container_width=True)
            with c2:
                st.plotly_chart(hbar_chart(fdf_et, "Equipment Type", "Avg Resolve Time by Equipment Type"), use_container_width=True)

    if optional_cols["Zone"]:
        with tab["By Zone"]:
            tab_pdf_download("By_Zone", fdf)
            fdf_z = col_filters(fdf, "byzone", cols_override=["Resolver","Shift","Andon Type","Equipment Type","Date"])
            st.markdown('<div class="sec-title">Count of Resolver and Avg Dwell Time by Creation Date and Zone</div>', unsafe_allow_html=True)
            tbl_z = build_group_pivot(fdf_z, "Zone")
            st.dataframe(apply_pivot_style(tbl_z), use_container_width=True, height=400)
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(donut_chart(fdf_z, "Zone", "Count of Records by Zone"), use_container_width=True)
            with c2:
                st.plotly_chart(hbar_chart(fdf_z, "Zone", "Avg Dwell Time by Zone"), use_container_width=True)

    if optional_cols["Shift"]:
        with tab["By Shift"]:
            tab_pdf_download("By_Shift", fdf)
            fdf_sh = col_filters(fdf, "byshift", cols_override=["Resolver","Zone","Andon Type","Equipment Type","Date"])
            st.markdown('<div class="sec-title">Count of Resolver and Avg Dwell Time by Creation Date and Shift</div>', unsafe_allow_html=True)
            tbl_sh = build_group_pivot(fdf_sh, "Shift")
            st.dataframe(apply_pivot_style(tbl_sh), use_container_width=True, height=400)
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(donut_chart(fdf_sh, "Shift", "Count of Resolver by Shift"), use_container_width=True)
            with c2:
                st.plotly_chart(hbar_chart(fdf_sh, "Shift", "Avg Dwell Time by Shift"), use_container_width=True)

    if optional_cols["Blocking"]:
        with tab["Blocking Analysis"]:
            tab_pdf_download("Blocking_Analysis", fdf)
            st.markdown('<div class="sec-title">Number of Andons and Resolution Times by Resolver and Blocking Status</div>', unsafe_allow_html=True)
            bl_vals = sorted(fdf["Blocking"].dropna().unique())
            bl_res_cols = {}
            for bv in bl_vals:
                sub_b = fdf[fdf["Blocking"] == bv]
                bl_res_cols[(str(bv), "Andons")]   = sub_b.groupby("Resolver")["Resolve_Min"].count()
                bl_res_cols[(str(bv), "Avg Time")] = sub_b.groupby("Resolver")["Resolve_Min"].mean().round(2)
            all_res = sorted(fdf["Resolver"].unique())
            bl_res_cols[("Total", "Andons")]   = fdf.groupby("Resolver")["Resolve_Min"].count().reindex(all_res, fill_value=0).astype(int)
            bl_res_cols[("Total", "Avg Time")] = fdf.groupby("Resolver")["Resolve_Min"].mean().round(2).reindex(all_res)
            bl_tbl = pd.DataFrame(bl_res_cols).reindex(all_res)
            bl_tbl.columns = pd.MultiIndex.from_tuples(bl_tbl.columns)
            bl_tbl.index.name = "Resolver"
            bl_grand = {}
            for bv in bl_vals:
                sub_b = fdf[fdf["Blocking"] == bv]
                bl_grand[(str(bv), "Andons")]   = int(sub_b["Resolve_Min"].count())
                bl_grand[(str(bv), "Avg Time")] = round(sub_b["Resolve_Min"].mean(), 2)
            bl_grand[("Total", "Andons")]   = int(fdf["Resolve_Min"].count())
            bl_grand[("Total", "Avg Time")] = round(fdf["Resolve_Min"].mean(), 2)
            bl_grand_row = pd.DataFrame(bl_grand, index=["Grand Total"])
            bl_grand_row.columns = pd.MultiIndex.from_tuples(bl_grand_row.columns)
            # Grand Total at BOTTOM
            bl_tbl = pd.concat([bl_tbl, bl_grand_row])
            bl_avg_cols_list = [(str(bv), "Avg Time") for bv in bl_vals] + [("Total", "Avg Time")]

            def _style_bl(data):
                s = pd.DataFrame("", index=data.index, columns=data.columns)
                rows = [i for i in data.index if i != "Grand Total"]
                for col in bl_avg_cols_list:
                    if col in data.columns:
                        ser = data.loc[rows, col]
                        for idx in rows:
                            s.loc[idx, col] = dwell_color(data.loc[idx, col], ser)
                s.loc["Grand Total"] = "font-weight:700; background-color:#e8eaf6; color:#1a237e"
                return s

            bl_styler = (bl_tbl.style.apply(_style_bl, axis=None)
                .format({c: "{:.2f}" for c in bl_tbl.columns if c[1] == "Avg Time"}, na_rep="—")
                .format({c: "{:,.0f}" for c in bl_tbl.columns if c[1] == "Andons"}, na_rep="—"))
            st.dataframe(bl_styler, use_container_width=True, height=420)
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(donut_chart(fdf, "Blocking", "Blocking vs Non-Blocking Distribution"), use_container_width=True)
            with c2:
                st.plotly_chart(hbar_chart(fdf, "Blocking", "Avg Resolve Time: Blocking vs Non-Blocking"), use_container_width=True)

    # ── NEW Tab: Department Blocking Analysis ─────────────────────────────────
    if optional_cols["Blocking"] and optional_cols["Equipment ID"]:
        with tab["🏭 Dept Blocking Analysis"]:

            def _norm_bl2(v):
                if pd.isna(v):
                    return "no"
                return "yes" if str(v).strip().lower() in ("yes", "true", "1", "y") else "no"

            dept_base = fdf.copy()
            dept_base["_BL"] = dept_base["Blocking"].apply(_norm_bl2)
            dept_base["Department"] = dept_base["Equipment ID"].apply(
                lambda x: get_department(str(x).strip()) if not pd.isna(x) else "Universal"
            )

            # Only blocking andons, exclude Product Problem & Out of Work
            dept_bl = dept_base[
                (dept_base["_BL"] == "yes") &
                (~dept_base["Andon Type"].isin(BLOCKING_EXCLUDE_TYPES))
            ].copy()

            st.markdown('<div class="sec-title">🏭 Department Blocking Analysis</div>', unsafe_allow_html=True)
            st.caption(
                "Blocking andons (Blocking = Yes) only · Excludes Product Problem & Out of Work · "
                "Stations mapped to ARSAW, PTR, ARStow, or Universal (all others)"
            )

            # ── Department filter ──────────────────────────────────────────────
            dept_opts = ["All"] + ["ARSAW", "PTR", "ARStow", "Universal"]
            dept_sel = st.selectbox("🏢 Filter by Department", dept_opts, key="dept_bl_sel")
            if dept_sel != "All":
                dept_bl_view = dept_bl[dept_bl["Department"] == dept_sel].copy()
            else:
                dept_bl_view = dept_bl.copy()

            if dept_bl_view.empty:
                st.warning("No blocking andons found for the selected department / current filters.")
            else:
                # ── KPI row ────────────────────────────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                dept_kpi_cols = st.columns(len(DEPT_COLORS) + 1)

                dept_summary = {}
                for dept, color in DEPT_COLORS.items():
                    sub = dept_bl[dept_bl["Department"] == dept]
                    cnt = len(sub)
                    hrs = sub["Resolve_Min"].sum() / 60
                    dept_summary[dept] = {"count": cnt, "hours": hrs, "color": color}

                total_hrs_all = dept_bl["Resolve_Min"].sum() / 60
                total_cnt_all = len(dept_bl)

                for i, (dept, info) in enumerate(dept_summary.items()):
                    dept_kpi_cols[i].markdown(f"""
                    <div class="kpi-box" style="border-top-color:{info['color']};">
                        <div class="kpi-label" style="color:{info['color']};">{dept}</div>
                        <div class="kpi-value" style="font-size:1.6rem;color:{info['color']};">{info['hours']:.2f} hrs</div>
                        <div class="kpi-sub">{info['count']:,} blocking andons</div>
                    </div>""", unsafe_allow_html=True)

                dept_kpi_cols[-1].markdown(f"""
                <div class="kpi-box" style="border-top-color:#9c27b0;">
                    <div class="kpi-label" style="color:#9c27b0;">TOTAL</div>
                    <div class="kpi-value" style="font-size:1.6rem;color:#9c27b0;">{total_hrs_all:.2f} hrs</div>
                    <div class="kpi-sub">{total_cnt_all:,} blocking andons combined</div>
                </div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Visual 1: Hours Lost per Department (donut) ────────────────
                vc1, vc2, vc3 = st.columns(3)

                with vc1:
                    st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>🕐 Hours Lost by Department</div>", unsafe_allow_html=True)
                    dept_hrs_df = pd.DataFrame([
                        {"Department": d, "Hours Lost": info["hours"], "Count": info["count"]}
                        for d, info in dept_summary.items() if info["count"] > 0
                    ])
                    fig_dept_donut = px.pie(
                        dept_hrs_df, names="Department", values="Hours Lost", hole=0.55,
                        color="Department",
                        color_discrete_map=DEPT_COLORS,
                    )
                    fig_dept_donut.update_traces(
                        textinfo="percent+label",
                        hovertemplate="<b>%{label}</b><br>Hours Lost: %{value:.2f}<br>%{percent}<extra></extra>"
                    )
                    fig_dept_donut.add_annotation(
                        text=f"<b>{total_hrs_all:.1f}h</b>", x=0.5, y=0.5,
                        font_size=18, font_color=_text, showarrow=False
                    )
                    fig_dept_donut.update_layout(
                        height=320, showlegend=True,
                        legend=dict(orientation="h", y=-0.3, font=dict(size=9)),
                        margin=dict(t=10, b=10, l=0, r=0),
                        paper_bgcolor="rgba(0,0,0,0)", font=dict(color=_text)
                    )
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.plotly_chart(fig_dept_donut, use_container_width=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                with vc2:
                    st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>📊 Blocking Count by Department</div>", unsafe_allow_html=True)
                    fig_dept_bar = go.Figure()
                    for dept, info in dept_summary.items():
                        fig_dept_bar.add_trace(go.Bar(
                            name=dept, x=[dept], y=[info["count"]],
                            marker_color=info["color"],
                            text=[info["count"]], textposition="outside",
                            hovertemplate=f"<b>{dept}</b><br>Count: {info['count']:,}<extra></extra>"
                        ))
                    fig_dept_bar.update_layout(
                        height=320, showlegend=False,
                        xaxis_title="", yaxis_title="Blocking Andons",
                        margin=dict(t=10, b=20, l=0, r=0),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color=_text),
                        xaxis=dict(color=_text), yaxis=dict(gridcolor="#333" if DM else "#eee", color=_text)
                    )
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.plotly_chart(fig_dept_bar, use_container_width=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                with vc3:
                    st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>⏱️ Avg Dwell Time by Department</div>", unsafe_allow_html=True)
                    dept_avg_df = []
                    for dept, info in dept_summary.items():
                        sub = dept_bl[dept_bl["Department"] == dept]
                        if len(sub) > 0:
                            dept_avg_df.append({"Department": dept, "Avg Time (min)": sub["Resolve_Min"].mean(), "color": info["color"]})
                    dept_avg_df = pd.DataFrame(dept_avg_df).sort_values("Avg Time (min)")
                    fig_dept_avg = go.Figure(go.Bar(
                        x=dept_avg_df["Avg Time (min)"].round(2),
                        y=dept_avg_df["Department"],
                        orientation="h",
                        marker_color=dept_avg_df["color"].tolist(),
                        text=dept_avg_df["Avg Time (min)"].round(2),
                        textposition="outside",
                    ))
                    fig_dept_avg.add_vline(x=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                                           annotation_text=f"Target ({DEFAULT_THRESHOLD}min)")
                    fig_dept_avg.update_layout(
                        height=320, xaxis_title="Avg Dwell Time (min)", yaxis_title="",
                        margin=dict(t=10, b=20, l=0, r=40),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color=_text),
                        xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
                        yaxis=dict(color=_text)
                    )
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.plotly_chart(fig_dept_avg, use_container_width=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Visual 2: Daily trend per department ───────────────────────
                st.markdown(f"<div style='font-size:0.85rem;font-weight:700;color:{_text};margin-bottom:4px;'>📈 Daily Hours Lost by Department</div>", unsafe_allow_html=True)
                daily_dept = (
                    dept_bl.groupby(["Date", "Department"])
                    .agg(Hours=("Resolve_Min", lambda x: x.sum()/60), Count=("Resolve_Min","count"))
                    .reset_index()
                )
                fig_daily_dept = px.bar(
                    daily_dept, x="Date", y="Hours", color="Department",
                    barmode="stack",
                    color_discrete_map=DEPT_COLORS,
                    labels={"Hours": "Hours Lost", "Date": ""},
                    hover_data={"Count": True}
                )
                fig_daily_dept.update_layout(
                    height=340, legend=dict(orientation="h", y=-0.25, font_size=10),
                    margin=dict(t=10, b=10, l=0, r=0),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_text),
                    xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
                    yaxis=dict(gridcolor="#333" if DM else "#eee", color=_text, title="Hours Lost"),
                )
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.plotly_chart(fig_daily_dept, use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Department sub-sections ────────────────────────────────────
                dept_list_show = [dept_sel] if dept_sel != "All" else list(DEPT_COLORS.keys())
                for dept in dept_list_show:
                    dept_color = DEPT_COLORS.get(dept, "#7986cb")
                    dept_data = dept_bl[dept_bl["Department"] == dept].copy()
                    if dept_data.empty:
                        continue

                    dept_hrs  = dept_data["Resolve_Min"].sum() / 60
                    dept_cnt  = len(dept_data)
                    dept_avg  = dept_data["Resolve_Min"].mean()

                    st.markdown(f"""
                    <div style="background:{'rgba(30,34,53,0.7)' if DM else '#f8f9fa'};
                                border-left:6px solid {dept_color}; border-radius:12px;
                                padding:1rem 1.4rem; margin:0.8rem 0 0.5rem 0;
                                box-shadow:{'0 2px 16px rgba(0,0,0,0.3)' if DM else '0 2px 10px rgba(0,0,0,0.08)'};">
                        <div style="font-size:1.2rem;font-weight:900;color:{dept_color};">🏭 {dept}</div>
                        <div style="display:flex;gap:2rem;margin-top:6px;flex-wrap:wrap;">
                            <span style="font-size:0.85rem;color:{_text};"><b style="color:{dept_color};">{dept_hrs:.2f} hrs</b> lost</span>
                            <span style="font-size:0.85rem;color:{_text};"><b style="color:{dept_color};">{dept_cnt:,}</b> blocking andons</span>
                            <span style="font-size:0.85rem;color:{_text};">Avg dwell: <b style="color:{dept_color};">{dept_avg:.2f} min</b></span>
                        </div>
                    </div>""", unsafe_allow_html=True)

                    da1, da2 = st.columns(2)

                    with da1:
                        # Andon type breakdown for this dept
                        st.markdown(f"<div style='font-size:0.8rem;font-weight:700;color:{_text};margin:6px 0 3px;'>Blocking Andon Types — {dept}</div>", unsafe_allow_html=True)
                        dept_type_summary = (
                            dept_data.groupby("Andon Type")
                            .agg(Count=("Resolve_Min","count"), Hrs_Lost=("Resolve_Min", lambda x: x.sum()/60), Avg_Min=("Resolve_Min","mean"))
                            .reset_index()
                            .sort_values("Hrs_Lost", ascending=False)
                        )
                        dept_type_summary["Hrs_Lost"] = dept_type_summary["Hrs_Lost"].round(2)
                        dept_type_summary["Avg_Min"]  = dept_type_summary["Avg_Min"].round(2)
                        dept_type_summary["% of Dept"] = (dept_type_summary["Count"] / dept_cnt * 100).round(1)
                        dept_type_summary.columns = ["Andon Type","Count","Hrs Lost","Avg (min)","% of Dept"]

                        def _style_dept_tbl(data):
                            s = pd.DataFrame("", index=data.index, columns=data.columns)
                            for idx in data.index:
                                avg = data.loc[idx, "Avg (min)"]
                                try:
                                    if avg > DEFAULT_THRESHOLD * 1.5:
                                        s.loc[idx, "Avg (min)"] = "background-color:rgb(210,40,40);color:white;font-weight:700"
                                    elif avg > DEFAULT_THRESHOLD:
                                        s.loc[idx, "Avg (min)"] = "background-color:rgb(255,140,0);color:black;font-weight:700"
                                    else:
                                        s.loc[idx, "Avg (min)"] = "background-color:rgb(60,180,60);color:white;font-weight:700"
                                except:
                                    pass
                            return s

                        st.dataframe(
                            dept_type_summary.style.apply(_style_dept_tbl, axis=None)
                            .format({"Count": "{:,}", "Hrs Lost": "{:.2f}", "Avg (min)": "{:.2f}", "% of Dept": "{:.1f}%"}),
                            use_container_width=True, hide_index=True, height=250
                        )

                    with da2:
                        # Top resolvers for this dept
                        st.markdown(f"<div style='font-size:0.8rem;font-weight:700;color:{_text};margin:6px 0 3px;'>Top Resolvers — {dept}</div>", unsafe_allow_html=True)
                        dept_res_summary = (
                            dept_data.groupby("Resolver")
                            .agg(Count=("Resolve_Min","count"), Hrs=("Resolve_Min", lambda x: x.sum()/60), Avg=("Resolve_Min","mean"))
                            .reset_index()
                            .sort_values("Count", ascending=False)
                            .head(10)
                        )
                        dept_res_summary["Hrs"] = dept_res_summary["Hrs"].round(2)
                        dept_res_summary["Avg"] = dept_res_summary["Avg"].round(2)
                        dept_res_summary.columns = ["Resolver","Count","Hrs Lost","Avg (min)"]
                        st.dataframe(
                            dept_res_summary.style.format({"Count": "{:,}", "Hrs Lost": "{:.2f}", "Avg (min)": "{:.2f}"}),
                            use_container_width=True, hide_index=True, height=250
                        )

                    # Detailed log: who created, when, who resolved
                    with st.expander(f"📋 Detailed Blocking Log — {dept}", expanded=False):
                        detail_cols = ["Time Created", "Andon Type", "Equipment ID", "Resolver", "Resolve_Min"]
                        if optional_cols["Creator"]:    detail_cols.insert(3, "Creator")
                        if optional_cols["Time Resolved"]: detail_cols.append("Time Resolved")
                        if "Zone" in dept_data.columns: detail_cols.append("Zone")

                        avail_detail = [c for c in detail_cols if c in dept_data.columns]
                        detail_display = dept_data[avail_detail].sort_values("Time Created", ascending=False).rename(
                            columns={"Resolve_Min": "Dwell (min)"}
                        )
                        st.dataframe(detail_display, use_container_width=True, height=350, hide_index=True)

                    st.markdown("<br>", unsafe_allow_html=True)

                # ── Combined summary table ─────────────────────────────────────
                st.markdown('<div class="sec-title">📊 Department Summary Table</div>', unsafe_allow_html=True)

                dept_summary_rows = []
                for dept in DEPT_COLORS.keys():
                    sub = dept_bl[dept_bl["Department"] == dept]
                    if sub.empty:
                        continue
                    # Top andon type
                    top_type = sub["Andon Type"].value_counts().index[0] if len(sub) > 0 else "—"
                    top_type_cnt = sub["Andon Type"].value_counts().iloc[0] if len(sub) > 0 else 0
                    dept_summary_rows.append({
                        "Department":       dept,
                        "Blocking Andons":  len(sub),
                        "Hours Lost":       round(sub["Resolve_Min"].sum()/60, 2),
                        "Avg Dwell (min)":  round(sub["Resolve_Min"].mean(), 2),
                        "Top Andon Type":   f"{top_type} ({top_type_cnt})",
                        "Unique Stations":  sub["Equipment ID"].nunique(),
                        "Unique Resolvers": sub["Resolver"].nunique(),
                    })

                # Grand total row
                dept_summary_rows.append({
                    "Department":       "GRAND TOTAL",
                    "Blocking Andons":  len(dept_bl),
                    "Hours Lost":       round(dept_bl["Resolve_Min"].sum()/60, 2),
                    "Avg Dwell (min)":  round(dept_bl["Resolve_Min"].mean(), 2),
                    "Top Andon Type":   f"{dept_bl['Andon Type'].value_counts().index[0]} ({dept_bl['Andon Type'].value_counts().iloc[0]})" if len(dept_bl) > 0 else "—",
                    "Unique Stations":  dept_bl["Equipment ID"].nunique(),
                    "Unique Resolvers": dept_bl["Resolver"].nunique(),
                })

                summary_df = pd.DataFrame(dept_summary_rows)

                def _style_summary(data):
                    s = pd.DataFrame("", index=data.index, columns=data.columns)
                    for idx in data.index:
                        dept = data.loc[idx, "Department"]
                        color = DEPT_COLORS.get(dept)
                        if dept == "GRAND TOTAL":
                            s.loc[idx] = "font-weight:700; background-color:#e8eaf6; color:#1a237e"
                        elif color:
                            avg = data.loc[idx, "Avg Dwell (min)"]
                            if avg >= DEFAULT_THRESHOLD * 1.5:
                                s.loc[idx, "Avg Dwell (min)"] = "background-color:rgb(210,40,40);color:white;font-weight:700"
                            elif avg >= DEFAULT_THRESHOLD:
                                s.loc[idx, "Avg Dwell (min)"] = "background-color:rgb(255,140,0);color:black;font-weight:700"
                            else:
                                s.loc[idx, "Avg Dwell (min)"] = "background-color:rgb(60,180,60);color:white;font-weight:700"
                    return s

                st.dataframe(
                    summary_df.style.apply(_style_summary, axis=None)
                    .format({"Blocking Andons": "{:,}", "Hours Lost": "{:.2f}", "Avg Dwell (min)": "{:.2f}"}),
                    use_container_width=True, hide_index=True
                )

                # ── Download buttons ───────────────────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                dl1, dl2 = st.columns(2)

                import io as _io_dept
                from openpyxl import Workbook as _WBd
                from openpyxl.styles import (PatternFill as _PFd, Font as _FNd,
                                              Alignment as _ALd, Border as _BDd, Side as _SDd)
                from openpyxl.utils import get_column_letter as _gcld

                def _build_dept_excel(dept_bl_df):
                    wb = _WBd()
                    THIN_d = _SDd(style="thin", color="C5CAE9")
                    BDR_d  = _BDd(left=THIN_d, right=THIN_d, top=THIN_d, bottom=THIN_d)
                    CTR_d  = _ALd(horizontal="center", vertical="center")
                    HDR_d  = _PFd("solid", fgColor="1A237E")
                    HFNT_d = _FNd(color="FFFFFF", bold=True, size=9)
                    NFT_d  = _FNd(size=9)
                    GRD_d  = _PFd("solid", fgColor="E8EAF6")
                    GFN_d  = _FNd(color="1A237E", bold=True, size=9)

                    dept_fills = {
                        "ARSAW":     _PFd("solid", fgColor="7986CB"),
                        "PTR":       _PFd("solid", fgColor="66BB6A"),
                        "ARStow":    _PFd("solid", fgColor="FFA726"),
                        "Universal": _PFd("solid", fgColor="EF5350"),
                    }

                    # Sheet 1: Summary
                    ws_sum = wb.active
                    ws_sum.title = "Dept Summary"
                    sum_hdrs = ["Department","Blocking Andons","Hours Lost","Avg Dwell (min)","Top Andon Type","Unique Stations","Unique Resolvers"]
                    for ci, h in enumerate(sum_hdrs, 1):
                        c = ws_sum.cell(row=1, column=ci, value=h)
                        c.fill=HDR_d; c.font=HFNT_d; c.alignment=CTR_d; c.border=BDR_d
                    for ri, row in enumerate(dept_summary_rows, 2):
                        for ci, key in enumerate(sum_hdrs, 1):
                            cell = ws_sum.cell(row=ri, column=ci, value=row.get(key))
                            cell.border = BDR_d; cell.alignment = CTR_d
                            if row["Department"] == "GRAND TOTAL":
                                cell.fill=GRD_d; cell.font=GFN_d
                            else:
                                fill = dept_fills.get(row["Department"])
                                if fill and ci == 1:
                                    cell.fill=fill; cell.font=_FNd(color="FFFFFF", bold=True, size=9)
                                else:
                                    cell.font=NFT_d
                    for col_obj in ws_sum.columns:
                        mx=0; cl=_gcld(col_obj[0].column)
                        for c in col_obj:
                            try: mx=max(mx, len(str(c.value or "")))
                            except: pass
                        ws_sum.column_dimensions[cl].width = min(mx+3, 35)

                    # One sheet per department
                    for dept in DEPT_COLORS.keys():
                        sub = dept_bl_df[dept_bl_df["Department"] == dept].copy()
                        if sub.empty:
                            continue
                        ws_d = wb.create_sheet(dept[:31])
                        det_cols = ["Time Created","Andon Type","Equipment ID","Resolver","Resolve_Min"]
                        if "Creator" in sub.columns:      det_cols.insert(3, "Creator")
                        if "Time Resolved" in sub.columns: det_cols.append("Time Resolved")
                        if "Zone" in sub.columns:          det_cols.append("Zone")
                        avail = [c for c in det_cols if c in sub.columns]
                        disp_names = {c: c for c in avail}
                        disp_names["Resolve_Min"] = "Dwell (min)"
                        for ci, col in enumerate(avail, 1):
                            cell = ws_d.cell(row=1, column=ci, value=disp_names[col])
                            cell.fill=HDR_d; cell.font=HFNT_d; cell.alignment=CTR_d; cell.border=BDR_d
                        for ri, (_, row) in enumerate(sub[avail].sort_values("Time Created", ascending=False).iterrows(), 2):
                            for ci, col in enumerate(avail, 1):
                                val = row[col]
                                if col == "Resolve_Min" and not pd.isna(val):
                                    val = round(float(val), 2)
                                elif col in ("Time Created","Time Resolved") and not pd.isna(val):
                                    val = str(val)
                                cell = ws_d.cell(row=ri, column=ci, value=val)
                                cell.border=BDR_d; cell.font=NFT_d
                        for col_obj in ws_d.columns:
                            mx=0; cl=_gcld(col_obj[0].column)
                            for c in col_obj:
                                try: mx=max(mx, len(str(c.value or "")))
                                except: pass
                            ws_d.column_dimensions[cl].width = min(mx+3, 32)

                    buf_d = _io_dept.BytesIO()
                    wb.save(buf_d)
                    buf_d.seek(0)
                    return buf_d.getvalue()

                def _build_dept_pdf(dept_bl_df):
                    try:
                        from reportlab.lib.pagesizes import A4, landscape
                        from reportlab.lib import colors as rl_colors
                        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                        Paragraph, Spacer, HRFlowable)
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.units import cm

                        buf_p = _io_dept.BytesIO()
                        doc = SimpleDocTemplate(buf_p, pagesize=landscape(A4),
                                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                                topMargin=1.5*cm, bottomMargin=1.5*cm)
                        styles = getSampleStyleSheet()
                        story_p = []

                        title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=rl_colors.HexColor("#1a237e"))
                        sub_style   = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=9,  textColor=rl_colors.gray)
                        h2_style    = ParagraphStyle("h2",    parent=styles["Heading2"], textColor=rl_colors.HexColor("#3949ab"))

                        story_p.append(Paragraph("LCY3 AFM Dashboard — Department Blocking Analysis", title_style))
                        story_p.append(Paragraph(f"Total blocking andons: {len(dept_bl_df):,} · Hours lost: {dept_bl_df['Resolve_Min'].sum()/60:.2f} hrs", sub_style))
                        story_p.append(Spacer(1, 12))

                        # Summary table
                        sum_data = [["Department","Blocking Andons","Hours Lost","Avg Dwell (min)","Unique Stations","Unique Resolvers"]]
                        for row in dept_summary_rows:
                            sum_data.append([
                                row["Department"], f"{row['Blocking Andons']:,}",
                                f"{row['Hours Lost']:.2f}", f"{row['Avg Dwell (min)']:.2f}",
                                str(row["Unique Stations"]), str(row["Unique Resolvers"])
                            ])
                        sum_tbl = Table(sum_data, colWidths=[3.5*cm,3.5*cm,3*cm,3.5*cm,3.5*cm,3.5*cm])
                        sum_tbl.setStyle(TableStyle([
                            ("BACKGROUND", (0,0), (-1,0), rl_colors.HexColor("#1a237e")),
                            ("TEXTCOLOR",  (0,0), (-1,0), rl_colors.white),
                            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                            ("ROWBACKGROUNDS", (0,1), (-1,-2), [rl_colors.HexColor("#f5f5f5"), rl_colors.white]),
                            ("BACKGROUND", (0,-1), (-1,-1), rl_colors.HexColor("#e8eaf6")),
                            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
                            ("GRID", (0,0), (-1,-1), 0.4, rl_colors.HexColor("#c5cae9")),
                            ("FONTSIZE", (0,0), (-1,-1), 9),
                            ("PADDING", (0,0), (-1,-1), 5),
                            ("ALIGN", (0,0), (-1,-1), "CENTER"),
                        ]))
                        story_p.append(sum_tbl)
                        story_p.append(Spacer(1, 16))

                        # Per department detail
                        for dept in DEPT_COLORS.keys():
                            sub = dept_bl_df[dept_bl_df["Department"] == dept]
                            if sub.empty:
                                continue
                            hex_color = DEPT_COLORS[dept]
                            hex_color_clean = hex_color[1:] if hex_color.startswith("#") else hex_color
                            story_p.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor(hex_color_clean)))
                            story_p.append(Paragraph(f"{dept} — {len(sub):,} blocking andons · {sub['Resolve_Min'].sum()/60:.2f} hrs lost", h2_style))
                            # Andon type breakdown
                            type_data = [["Andon Type","Count","Hrs Lost","Avg (min)","% of Dept"]]
                            for _, row in (sub.groupby("Andon Type")
                                           .agg(Count=("Resolve_Min","count"), Hrs=("Resolve_Min",lambda x:x.sum()/60), Avg=("Resolve_Min","mean"))
                                           .reset_index().sort_values("Hrs", ascending=False)).iterrows():
                                type_data.append([
                                    str(row["Andon Type"])[:35], str(int(row["Count"])),
                                    f"{row['Hrs']:.2f}", f"{row['Avg']:.2f}",
                                    f"{row['Count']/len(sub)*100:.1f}%"
                                ])
                            t_tbl = Table(type_data, colWidths=[7*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm])
                            t_tbl.setStyle(TableStyle([
                                ("BACKGROUND", (0,0), (-1,0), rl_colors.HexColor("#3949ab")),
                                ("TEXTCOLOR",  (0,0), (-1,0), rl_colors.white),
                                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                                ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.HexColor("#f5f5f5"), rl_colors.white]),
                                ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#c5cae9")),
                                ("FONTSIZE", (0,0), (-1,-1), 8),
                                ("PADDING", (0,0), (-1,-1), 4),
                            ]))
                            story_p.append(t_tbl)
                            story_p.append(Spacer(1, 10))

                        doc.build(story_p)
                        buf_p.seek(0)
                        return buf_p.getvalue()
                    except ImportError:
                        return None

                with dl1:
                    dept_excel = _build_dept_excel(dept_bl)
                    st.download_button(
                        "⬇️ Download Dept Blocking (.xlsx)",
                        dept_excel, "Dept_Blocking_Analysis.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="dept_excel_dl"
                    )
                with dl2:
                    dept_pdf = _build_dept_pdf(dept_bl)
                    if dept_pdf:
                        st.download_button(
                            "⬇️ Download Dept Blocking (.pdf)",
                            dept_pdf, "Dept_Blocking_Analysis.pdf",
                            "application/pdf",
                            use_container_width=True, key="dept_pdf_dl"
                        )
                    else:
                        st.info("💡 Install `reportlab` to enable PDF download.")

    if optional_cols["Equipment ID"]:
        with tab["Equipment ID Analysis"]:
            tab_pdf_download("Equipment_ID_Analysis", fdf)

            # ── Per-tab column filters ────────────────────────────────────────
            fdf_eid = col_filters(fdf, "eid", cols_override=["Andon Type","Resolver","Zone","Shift","Date"])

            st.markdown('<div class="sec-title">🔧 Equipment ID — Detailed Analysis</div>', unsafe_allow_html=True)

            if fdf_eid.empty or "Equipment ID" not in fdf_eid.columns:
                st.warning("No Equipment ID data available.")
            else:
                eid_weeks_all = sorted(fdf_eid["Week"].dropna().unique(), reverse=True)

                # ── EID-level search / filter ─────────────────────────────────
                eid_search_col, eid_dept_col = st.columns([2, 2])
                with eid_search_col:
                    eid_search = st.text_input("🔎 Search Equipment ID", "", key="eid_search",
                                               placeholder="Type an ID, e.g. 2319")
                with eid_dept_col:
                    dept_opts_eid = ["All Departments", "ARSAW", "PTR", "ARStow", "Universal"]
                    eid_dept_filter = st.selectbox("🏭 Filter by Department", dept_opts_eid, key="eid_dept")

                # Apply EID-level filters
                fdf_eid_view = fdf_eid.copy()
                fdf_eid_view["Department"] = fdf_eid_view["Equipment ID"].apply(
                    lambda x: get_department(str(x).strip()) if not pd.isna(x) else "Universal"
                )
                if eid_search.strip():
                    fdf_eid_view = fdf_eid_view[
                        fdf_eid_view["Equipment ID"].astype(str).str.contains(eid_search.strip(), case=False, na=False)
                    ]
                if eid_dept_filter != "All Departments":
                    fdf_eid_view = fdf_eid_view[fdf_eid_view["Department"] == eid_dept_filter]

                if fdf_eid_view.empty:
                    st.warning("No Equipment IDs match the current filter.")
                else:
                    # ── KPI strip ─────────────────────────────────────────────
                    ek1, ek2, ek3, ek4, ek5 = st.columns(5)
                    for eb, elbl, eval_, esub in [
                        (ek1, "Unique Equipment IDs", f"{fdf_eid_view['Equipment ID'].nunique():,}", "in filtered view"),
                        (ek2, "Total Andons",          f"{len(fdf_eid_view):,}",                     "on these IDs"),
                        (ek3, "Avg Dwell Time",        f"{fdf_eid_view['Resolve_Min'].mean():.2f} min", "per andon"),
                        (ek4, "Worst ID Avg",
                              fdf_eid_view.groupby("Equipment ID")["Resolve_Min"].mean().idxmax()
                              if len(fdf_eid_view) > 0 else "—",
                              f"{fdf_eid_view.groupby('Equipment ID')['Resolve_Min'].mean().max():.2f} min avg" if len(fdf_eid_view) > 0 else ""),
                        (ek5, "Most Active ID",
                              fdf_eid_view.groupby("Equipment ID")["Resolve_Min"].count().idxmax()
                              if len(fdf_eid_view) > 0 else "—",
                              f"{fdf_eid_view.groupby('Equipment ID')['Resolve_Min'].count().max()} andons" if len(fdf_eid_view) > 0 else ""),
                    ]:
                        eb.markdown(f"""
                        <div class="kpi-box">
                            <div class="kpi-label">{elbl}</div>
                            <div class="kpi-value" style="font-size:1.4rem;">{eval_}</div>
                            <div class="kpi-sub">{esub}</div>
                        </div>""", unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)

                    # ── Week × Equipment ID pivot ─────────────────────────────
                    st.markdown('<div class="sec-title">📅 Andons by Equipment ID × Week</div>', unsafe_allow_html=True)
                    eid_count_p = fdf_eid_view.pivot_table(
                        index="Equipment ID", columns="Week",
                        values="Resolve_Min", aggfunc="count", fill_value=0
                    )
                    eid_avg_p = fdf_eid_view.pivot_table(
                        index="Equipment ID", columns="Week",
                        values="Resolve_Min", aggfunc="mean"
                    )
                    eid_wk_cols = {}
                    for w in eid_weeks_all:
                        if w in eid_count_p.columns:
                            eid_wk_cols[(f"Wk {w}", "Count")]    = eid_count_p[w].astype(int)
                            eid_wk_cols[(f"Wk {w}", "Avg (min)")] = eid_avg_p[w].round(2) if w in eid_avg_p.columns else pd.Series(dtype=float)
                    eid_wk_cols[("Total", "Count")]    = eid_count_p[eid_weeks_all].sum(axis=1).astype(int)
                    eid_wk_cols[("Total", "Avg (min)")] = fdf_eid_view.groupby("Equipment ID")["Resolve_Min"].mean().round(2)

                    eid_wk_tbl = pd.DataFrame(eid_wk_cols).sort_values(("Total","Count"), ascending=False)
                    eid_wk_tbl.index.name = "Equipment ID"

                    # Add department column
                    dept_map = fdf_eid_view.groupby("Equipment ID")["Department"].first()
                    eid_wk_tbl.insert(0, "Dept", eid_wk_tbl.index.map(dept_map))

                    # Grand Total row
                    eid_gt = {"Dept": "ALL"}
                    for w in eid_weeks_all:
                        if w in eid_count_p.columns:
                            sw = fdf_eid_view[fdf_eid_view["Week"] == w]
                            eid_gt[(f"Wk {w}", "Count")]     = int(sw["Resolve_Min"].count())
                            eid_gt[(f"Wk {w}", "Avg (min)")] = round(sw["Resolve_Min"].mean(), 2)
                    eid_gt[("Total","Count")]     = len(fdf_eid_view)
                    eid_gt[("Total","Avg (min)")] = round(fdf_eid_view["Resolve_Min"].mean(), 2)
                    eid_gt_row = pd.DataFrame(eid_gt, index=["Grand Total"])
                    eid_wk_tbl.columns = pd.MultiIndex.from_tuples(
                        [("Info","Dept")] + list(eid_wk_tbl.columns[1:])
                    )
                    eid_gt_row.columns = pd.MultiIndex.from_tuples(
                        [("Info","Dept")] + [(f"Wk {w}", c) for w in eid_weeks_all for c in ["Count","Avg (min)"] if w in eid_count_p.columns] +
                        [("Total","Count"), ("Total","Avg (min)")]
                    ) if len(eid_gt_row.columns) == len(eid_wk_tbl.columns) else eid_wk_tbl.columns[:len(eid_gt_row.columns)]

                    try:
                        eid_wk_full = pd.concat([eid_wk_tbl, eid_gt_row])
                    except Exception:
                        eid_wk_full = eid_wk_tbl.copy()

                    avg_cols_eid = [c for c in eid_wk_full.columns if "Avg" in str(c[1])]
                    cnt_cols_eid = [c for c in eid_wk_full.columns if "Count" in str(c[1])]

                    def _style_eid_wk(data):
                        s = pd.DataFrame("", index=data.index, columns=data.columns)
                        dr = [i for i in data.index if i != "Grand Total"]
                        for col in avg_cols_eid:
                            if col in data.columns:
                                ser = data.loc[dr, col]
                                for idx in dr:
                                    s.loc[idx, col] = dwell_color(data.loc[idx, col], ser)
                        # Dept colour coding
                        dept_clrs = {"ARSAW":"#7986cb","PTR":"#66bb6a","ARStow":"#ffa726","Universal":"#ef5350","ALL":"#9c27b0"}
                        for idx in dr:
                            try:
                                d = str(data.loc[idx, ("Info","Dept")])
                                clr = dept_clrs.get(d, "")
                                if clr:
                                    s.loc[idx, ("Info","Dept")] = f"font-weight:700;color:{clr}"
                            except Exception:
                                pass
                        if "Grand Total" in data.index:
                            s.loc["Grand Total"] = "font-weight:700;background-color:#e8eaf6;color:#1a237e"
                        return s

                    fmt_eid = {c: "{:.2f}" for c in avg_cols_eid}
                    fmt_eid.update({c: "{:,.0f}" for c in cnt_cols_eid})
                    st.dataframe(
                        eid_wk_full.style.apply(_style_eid_wk, axis=None).format(fmt_eid, na_rep="—"),
                        use_container_width=True, height=420
                    )

                    st.markdown("<br>", unsafe_allow_html=True)

                    # ── Charts row 1: Top IDs ─────────────────────────────────
                    st.markdown('<div class="sec-title">📊 Visual Breakdown</div>', unsafe_allow_html=True)
                    ec1, ec2 = st.columns(2)

                    with ec1:
                        st.markdown(f"<div style='font-size:0.82rem;font-weight:700;color:{_text};margin-bottom:4px;'>Top 20 Equipment IDs — Andon Count</div>", unsafe_allow_html=True)
                        top20_eid = (fdf_eid_view.groupby("Equipment ID")["Resolve_Min"].count()
                                     .nlargest(20).reset_index().rename(columns={"Resolve_Min": "Count"})
                                     .sort_values("Count"))
                        top20_eid["Dept"] = top20_eid["Equipment ID"].apply(lambda x: get_department(str(x)))
                        fig_eid_bar2 = px.bar(
                            top20_eid, x="Count", y="Equipment ID", orientation="h",
                            color="Dept", color_discrete_map=DEPT_COLORS,
                            text="Count"
                        )
                        fig_eid_bar2.update_traces(textposition="outside")
                        fig_eid_bar2.update_layout(
                            height=460, yaxis_title="", xaxis_title="Andon Count",
                            margin=dict(t=10, b=10, l=0, r=40),
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=_text),
                            xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
                            yaxis=dict(color=_text),
                            legend=dict(orientation="h", y=-0.15, font_size=10)
                        )
                        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                        st.plotly_chart(fig_eid_bar2, use_container_width=True)
                        st.markdown('</div>', unsafe_allow_html=True)

                    with ec2:
                        st.markdown(f"<div style='font-size:0.82rem;font-weight:700;color:{_text};margin-bottom:4px;'>Top 20 Equipment IDs — Avg Dwell Time</div>", unsafe_allow_html=True)
                        top20_avg2 = (fdf_eid_view.groupby("Equipment ID")
                                      .agg(Count=("Resolve_Min","count"), Avg=("Resolve_Min","mean"))
                                      .reset_index().nlargest(20, "Count")
                                      .sort_values("Avg", ascending=True))
                        top20_avg2["Dept"] = top20_avg2["Equipment ID"].apply(lambda x: get_department(str(x)))
                        top20_avg2["Color"] = top20_avg2["Avg"].apply(
                            lambda x: "rgb(210,40,40)" if x > DEFAULT_THRESHOLD*1.5
                            else "rgb(255,140,0)" if x > DEFAULT_THRESHOLD else "rgb(60,180,60)"
                        )
                        fig_eid_avg2 = go.Figure(go.Bar(
                            x=top20_avg2["Avg"].round(2),
                            y=top20_avg2["Equipment ID"].astype(str),
                            orientation="h",
                            marker_color=top20_avg2["Color"].tolist(),
                            text=top20_avg2["Avg"].round(2),
                            textposition="outside",
                            customdata=top20_avg2["Count"],
                            hovertemplate="<b>%{y}</b><br>Avg: %{x:.2f} min<br>Andons: %{customdata}<extra></extra>"
                        ))
                        fig_eid_avg2.add_vline(x=DEFAULT_THRESHOLD, line_dash="dash", line_color="gray",
                                               annotation_text=f"Target ({DEFAULT_THRESHOLD} min)")
                        fig_eid_avg2.update_layout(
                            height=460, xaxis_title="Avg Dwell Time (min)", yaxis_title="",
                            margin=dict(t=10, b=10, l=0, r=40),
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=_text),
                            xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text),
                            yaxis=dict(color=_text)
                        )
                        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                        st.plotly_chart(fig_eid_avg2, use_container_width=True)
                        st.markdown('</div>', unsafe_allow_html=True)

                    # ── Charts row 2: Distribution & Trend ────────────────────
                    ec3, ec4 = st.columns(2)

                    with ec3:
                        st.markdown(f"<div style='font-size:0.82rem;font-weight:700;color:{_text};margin:1rem 0 4px;'>🥧 Distribution by Department</div>", unsafe_allow_html=True)
                        dept_dist = fdf_eid_view.groupby("Department")["Resolve_Min"].count().reset_index()
                        dept_dist.columns = ["Department","Count"]
                        fig_dept_dist = px.pie(
                            dept_dist, names="Department", values="Count", hole=0.55,
                            color="Department", color_discrete_map=DEPT_COLORS
                        )
                        fig_dept_dist.update_traces(textinfo="percent+label")
                        fig_dept_dist.add_annotation(
                            text=f"<b>{dept_dist['Count'].sum():,}</b>",
                            x=0.5, y=0.5, font_size=20, font_color=_text, showarrow=False
                        )
                        fig_dept_dist.update_layout(
                            height=340, showlegend=False,
                            margin=dict(t=10, b=10, l=0, r=0),
                            paper_bgcolor="rgba(0,0,0,0)", font=dict(color=_text)
                        )
                        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                        st.plotly_chart(fig_dept_dist, use_container_width=True)
                        st.markdown('</div>', unsafe_allow_html=True)

                    with ec4:
                        st.markdown(f"<div style='font-size:0.82rem;font-weight:700;color:{_text};margin:1rem 0 4px;'>📈 Daily Andon Trend for Equipment IDs</div>", unsafe_allow_html=True)
                        eid_daily = fdf_eid_view.groupby("Date").agg(
                            Count=("Resolve_Min","count"),
                            Avg=("Resolve_Min","mean")
                        ).reset_index()
                        fig_eid_trend = go.Figure()
                        fig_eid_trend.add_trace(go.Bar(
                            x=eid_daily["Date"], y=eid_daily["Count"],
                            name="Daily Andons", marker_color=_accent,
                            opacity=0.7
                        ))
                        fig_eid_trend.add_trace(go.Scatter(
                            x=eid_daily["Date"], y=eid_daily["Avg"].round(2),
                            name="Avg Time (min)", yaxis="y2",
                            line=dict(color="#ffa726", width=2, dash="dot"),
                            marker=dict(size=5)
                        ))
                        fig_eid_trend.add_hline(
                            y=DEFAULT_THRESHOLD, line_dash="dash", line_color="#ef5350",
                            annotation_text=f"Target ({DEFAULT_THRESHOLD}min)", yref="y2"
                        )
                        fig_eid_trend.update_layout(
                            height=340,
                            yaxis=dict(title="Count", gridcolor="#333" if DM else "#eee", color=_text),
                            yaxis2=dict(title="Avg (min)", overlaying="y", side="right", showgrid=False, color="#ffa726"),
                            legend=dict(orientation="h", y=-0.2, font_size=10),
                            margin=dict(t=10, b=20, l=0, r=60),
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=_text),
                            xaxis=dict(gridcolor="#333" if DM else "#eee", color=_text)
                        )
                        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                        st.plotly_chart(fig_eid_trend, use_container_width=True)
                        st.markdown('</div>', unsafe_allow_html=True)

                    # ── Andon type breakdown per EID ──────────────────────────
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown('<div class="sec-title">🔍 Andon Type Breakdown by Equipment ID</div>', unsafe_allow_html=True)

                    eid_type_pivot = fdf_eid_view.pivot_table(
                        index="Equipment ID", columns="Andon Type",
                        values="Resolve_Min", aggfunc="count", fill_value=0
                    ).astype(int)
                    eid_type_pivot["Total"] = eid_type_pivot.sum(axis=1)
                    eid_type_pivot["Avg Dwell (min)"] = fdf_eid_view.groupby("Equipment ID")["Resolve_Min"].mean().round(2)
                    eid_type_pivot["Department"] = eid_type_pivot.index.map(
                        lambda x: get_department(str(x))
                    )
                    # Reorder cols: Dept first
                    cols_order = ["Department","Total","Avg Dwell (min)"] + [c for c in eid_type_pivot.columns if c not in ["Department","Total","Avg Dwell (min)"]]
                    eid_type_pivot = eid_type_pivot[cols_order].sort_values("Total", ascending=False)

                    def _style_eid_type(data):
                        s = pd.DataFrame("", index=data.index, columns=data.columns)
                        dept_clrs2 = {"ARSAW":"color:#7986cb;font-weight:700","PTR":"color:#66bb6a;font-weight:700",
                                      "ARStow":"color:#ffa726;font-weight:700","Universal":"color:#ef5350;font-weight:700"}
                        for idx in data.index:
                            try:
                                d = str(data.loc[idx,"Department"])
                                if d in dept_clrs2:
                                    s.loc[idx,"Department"] = dept_clrs2[d]
                            except Exception:
                                pass
                            try:
                                avg = float(data.loc[idx,"Avg Dwell (min)"])
                                if avg > DEFAULT_THRESHOLD*1.5:
                                    s.loc[idx,"Avg Dwell (min)"] = "background-color:rgb(210,40,40);color:white;font-weight:700"
                                elif avg > DEFAULT_THRESHOLD:
                                    s.loc[idx,"Avg Dwell (min)"] = "background-color:rgb(255,140,0);color:black;font-weight:700"
                                else:
                                    s.loc[idx,"Avg Dwell (min)"] = "background-color:rgb(60,180,60);color:white;font-weight:700"
                            except Exception:
                                pass
                        return s

                    st.dataframe(
                        eid_type_pivot.style.apply(_style_eid_type, axis=None)
                        .format({"Avg Dwell (min)": "{:.2f}", "Total": "{:,}"}, na_rep="—"),
                        use_container_width=True, height=400
                    )

                    # ── Single EID drill-down ─────────────────────────────────
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown('<div class="sec-title">🔬 Single Equipment ID — Full Drill-Down</div>', unsafe_allow_html=True)
                    all_eids = sorted(fdf_eid_view["Equipment ID"].dropna().astype(str).unique().tolist())
                    if all_eids:
                        drill_col1, drill_col2 = st.columns([2,2])
                        with drill_col1:
                            sel_eid = st.selectbox("Select Equipment ID", all_eids, key="eid_drill_sel")
                        eid_sub = fdf_eid_view[fdf_eid_view["Equipment ID"].astype(str) == sel_eid].copy()
                        eid_dept_val = get_department(sel_eid)
                        eid_dept_color = DEPT_COLORS.get(eid_dept_val, "#7986cb")

                        with drill_col2:
                            st.markdown(f"""
                            <div style="background:{'rgba(30,34,53,0.8)' if DM else '#f8f9fa'};
                                        border-left:5px solid {eid_dept_color};border-radius:10px;
                                        padding:10px 16px;margin-top:8px;">
                                <span style="font-size:1.1rem;font-weight:900;color:{eid_dept_color};">
                                    Station {sel_eid}
                                </span>
                                <span style="font-size:0.8rem;color:{_sub};margin-left:10px;">{eid_dept_val}</span>
                                <div style="font-size:0.82rem;color:{_text};margin-top:4px;">
                                    <b style="color:{eid_dept_color};">{len(eid_sub):,}</b> andons &nbsp;·&nbsp;
                                    <b style="color:{eid_dept_color};">{eid_sub["Resolve_Min"].mean():.2f} min</b> avg dwell
                                </div>
                            </div>""", unsafe_allow_html=True)

                        dd1, dd2 = st.columns(2)
                        with dd1:
                            # Timeline
                            eid_daily_sub = eid_sub.groupby("Date").agg(
                                Count=("Resolve_Min","count"), Avg=("Resolve_Min","mean")
                            ).reset_index()
                            fig_eid_dd = go.Figure()
                            fig_eid_dd.add_trace(go.Bar(
                                x=eid_daily_sub["Date"].astype(str), y=eid_daily_sub["Count"],
                                marker_color=eid_dept_color, opacity=0.75, name="Count"
                            ))
                            fig_eid_dd.add_trace(go.Scatter(
                                x=eid_daily_sub["Date"].astype(str), y=eid_daily_sub["Avg"].round(2),
                                yaxis="y2", name="Avg (min)",
                                line=dict(color="#ffa726", width=2),
                                marker=dict(size=6)
                            ))
                            fig_eid_dd.update_layout(
                                title=f"Daily trend — Station {sel_eid}",
                                height=280,
                                yaxis=dict(title="Count"),
                                yaxis2=dict(title="Avg (min)", overlaying="y", side="right", showgrid=False),
                                margin=dict(t=40,b=20,l=0,r=50),
                                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=_text),
                                legend=dict(orientation="h", y=-0.25)
                            )
                            st.plotly_chart(fig_eid_dd, use_container_width=True)

                        with dd2:
                            # Andon type breakdown for this EID
                            eid_type_sub = eid_sub["Andon Type"].value_counts().reset_index()
                            eid_type_sub.columns = ["Andon Type","Count"]
                            fig_eid_type_dd = px.pie(
                                eid_type_sub, names="Andon Type", values="Count", hole=0.55,
                                title=f"Andon Types — Station {sel_eid}",
                                color_discrete_sequence=px.colors.qualitative.Set2
                            )
                            fig_eid_type_dd.update_traces(textinfo="percent+label", textfont_size=9)
                            fig_eid_type_dd.update_layout(
                                height=280, showlegend=True,
                                legend=dict(orientation="h", y=-0.35, font_size=9),
                                margin=dict(t=40,b=10,l=0,r=0),
                                paper_bgcolor="rgba(0,0,0,0)", font=dict(color=_text)
                            )
                            st.plotly_chart(fig_eid_type_dd, use_container_width=True)

                        # Full log for this EID
                        with st.expander(f"📋 Full Andon Log — Station {sel_eid} ({len(eid_sub):,} records)", expanded=False):
                            log_cols = ["Time Created","Andon Type","Resolve_Min","Resolver"]
                            if "Creator"      in eid_sub.columns: log_cols.insert(3,"Creator")
                            if "Blocking"     in eid_sub.columns: log_cols.append("Blocking")
                            if "Zone"         in eid_sub.columns: log_cols.append("Zone")
                            if "Time Resolved" in eid_sub.columns: log_cols.append("Time Resolved")
                            avail_log = [c for c in log_cols if c in eid_sub.columns]
                            st.dataframe(
                                eid_sub[avail_log].sort_values("Time Created", ascending=False)
                                .rename(columns={"Resolve_Min":"Dwell (min)"}),
                                use_container_width=True, height=320, hide_index=True
                            )

                    # ── Download ──────────────────────────────────────────────
                    st.markdown("<br>", unsafe_allow_html=True)
                    import io as _io_eid
                    eid_csv = fdf_eid_view.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        "⬇️ Download Equipment ID Data (CSV)",
                        eid_csv, "Equipment_ID_Analysis.csv", "text/csv",
                        use_container_width=True, key="eid_dl_csv"
                    )

    # ── Tab: Hourly Trend
    # ── Tab: Hourly Trend ─────────────────────────────────────────────────────
    with tab["Hourly Trend"]:
        tab_pdf_download("Hourly_Trend", fdf)
        fdf_hr = col_filters(fdf, "hourly", cols_override=["Andon Type","Resolver","Zone","Shift","Date"])
        st.markdown('<div class="sec-title">Count of Andon Type and Avg Dwell Time by Hour of Day</div>', unsafe_allow_html=True)
        hourly_count = (fdf_hr.groupby(["Hour", "Andon Type"])["Resolve_Min"]
                        .count().reset_index().rename(columns={"Resolve_Min": "Count"}))
        hourly_avg   = (fdf_hr.groupby("Hour")["Resolve_Min"]
                        .mean().reset_index().rename(columns={"Resolve_Min": "Avg Time"}))
        fig_h = go.Figure()
        colors = px.colors.qualitative.Pastel + px.colors.qualitative.Set2
        for i, at in enumerate(sorted(fdf_hr["Andon Type"].dropna().unique())):
            sub = hourly_count[hourly_count["Andon Type"] == at]
            fig_h.add_trace(go.Bar(x=sub["Hour"], y=sub["Count"], name=at,
                                   marker_color=colors[i % len(colors)]))
        fig_h.add_trace(go.Scatter(
            x=hourly_avg["Hour"], y=hourly_avg["Avg Time"],
            name="Avg Dwell Time", yaxis="y2", mode="lines+markers",
            line=dict(color="#b71c1c", width=2.5), marker=dict(size=6, color="#b71c1c")
        ))
        fig_h.update_layout(barmode="stack", height=500,
                            xaxis=dict(title="Hour of Day", dtick=1),
                            yaxis=dict(title="Andon Count"),
                            yaxis2=dict(title="Avg Dwell Time (min)", overlaying="y", side="right", showgrid=False),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font_size=10),
                            margin=dict(t=80, b=40, l=0, r=0))
        st.plotly_chart(fig_h, use_container_width=True)

        st.markdown('<div class="sec-title">Resolve Time Trend by Day</div>', unsafe_allow_html=True)
        daily = (fdf_hr.groupby("Date").agg(Count=("Resolve_Min", "count"), Avg=("Resolve_Min", "mean"))
                 .reset_index().sort_values("Date"))
        fig_d = go.Figure()
        fig_d.add_trace(go.Bar(x=daily["Date"], y=daily["Count"], name="Andons",
                               marker_color="#90caf9", yaxis="y"))
        fig_d.add_trace(go.Scatter(x=daily["Date"], y=daily["Avg"].round(2),
                                   name="Avg Dwell Time", yaxis="y2", mode="lines+markers",
                                   line=dict(color="#1a237e", width=2.5)))
        fig_d.update_layout(barmode="group", height=380,
                            xaxis=dict(title="Date"), yaxis=dict(title="Count"),
                            yaxis2=dict(title="Avg Dwell Time (min)", overlaying="y", side="right", showgrid=False),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02),
                            margin=dict(t=60, b=40, l=0, r=0))
        st.plotly_chart(fig_d, use_container_width=True)

    # ── Tab: Heatmap ──────────────────────────────────────────────────────────
    with tab["Heatmap"]:
        tab_pdf_download("Heatmap", fdf)
        fdf_hm = col_filters(fdf, "heatmap", cols_override=["Zone","Shift","Date"])
        st.markdown('<div class="sec-title">Avg Resolve Time Heatmap — Resolver × Andon Type</div>', unsafe_allow_html=True)
        pivot_hm = fdf_hm.pivot_table(index="Resolver", columns="Andon Type",
                                   values="Resolve_Min", aggfunc="mean").round(2)
        fig_hm = px.imshow(pivot_hm, text_auto=True, aspect="auto",
                           color_continuous_scale="RdYlGn_r", labels=dict(color="Avg (min)"))
        fig_hm.update_layout(height=max(350, len(pivot_hm) * 40 + 100), margin=dict(t=30, b=40, l=0, r=0))
        st.plotly_chart(fig_hm, use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            res_perf = (fdf_hm.groupby("Resolver")
                        .agg(Count=("Resolve_Min", "count"), Avg=("Resolve_Min", "mean"))
                        .reset_index().sort_values("Avg", ascending=False))
            fig_r = px.bar(res_perf, x="Resolver", y="Avg", color="Avg",
                           color_continuous_scale="RdYlGn_r", text=res_perf["Avg"].round(2),
                           labels={"Avg": "Avg Time (min)"})
            fig_r.update_traces(textposition="outside")
            fig_r.update_layout(coloraxis_showscale=False, height=380,
                                xaxis_title="", yaxis_title="Avg Dwell Time (min)",
                                margin=dict(t=20, b=40, l=0, r=0))
            st.plotly_chart(fig_r, use_container_width=True)
        with c2:
            cat_perf = (fdf_hm.groupby("Andon Type")
                        .agg(Count=("Resolve_Min", "count"), Avg=("Resolve_Min", "mean"))
                        .reset_index().sort_values("Avg"))
            fig_c = px.bar(cat_perf, x="Avg", y="Andon Type", orientation="h",
                           color="Avg", color_continuous_scale="Blues", text=cat_perf["Avg"].round(2),
                           labels={"Avg": "Avg Time (min)"})
            fig_c.update_traces(textposition="outside")
            fig_c.update_layout(coloraxis_showscale=False, height=380,
                                yaxis_title="", xaxis_title="Avg Dwell Time (min)",
                                margin=dict(t=20, b=40, l=0, r=0))
            st.plotly_chart(fig_c, use_container_width=True)

    # ── Tab: Raw Data ─────────────────────────────────────────────────────────
    with tab["Raw Data"]:
        tab_pdf_download("Raw_Data", fdf)
        fdf_raw = col_filters(fdf, "rawdata")
        st.markdown('<div class="sec-title">Raw Resolved Andon Records</div>', unsafe_allow_html=True)
        st.markdown(f"**{len(fdf_raw):,}** records matching current filters · from **{len(uploaded_files)}** file(s)")
        st.dataframe(fdf_raw, use_container_width=True, height=500)
        dl1, dl2 = st.columns(2)
        with dl1:
            csv = fdf_raw.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ Download as CSV", csv, "andon_filtered.csv", "text/csv", use_container_width=True)
        with dl2:
            import io
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                fdf.to_excel(writer, index=False, sheet_name="Andon Data")
                lb_export = lb[["Resolver", "Total_Andons", "Avg_Time", "Efficiency", "Within_Threshold", "Badge", "Status"]].copy()
                lb_export.columns = ["Resolver", "Total Andons", "Avg Time (min)", "Efficiency Score", "% Within Threshold", "Badge", "Status"]
                lb_export.to_excel(writer, index=True, index_label="Rank", sheet_name="Leaderboard")
            buf.seek(0)
            st.download_button("⬇️ Download as Excel", buf.getvalue(), "andon_report.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    # ── Tab: Export ───────────────────────────────────────────────────────────
    with tab["📤 Export"]:
        st.markdown('<div class="sec-title">Export Full Reports</div>', unsafe_allow_html=True)
        st.markdown(
            "Download reports as **Excel (.xlsx)** or **PDF (.pdf)** — "
            "each contains summary KPIs, pivot tables, leaderboard and andon type breakdown."
        )
        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown('<div class="sec-title">📅 Daily Reports</div>', unsafe_allow_html=True)
        d1, d2 = st.columns(2)

        with d1:
            st.markdown("""
            <div style="background:#e8f5e9; border-left:5px solid #388e3c; border-radius:8px;
                        padding:16px 20px; margin-bottom:12px;">
                <div style="font-size:1.1rem; font-weight:700; color:#1b5e20;">📊 Daily Excel Report</div>
                <div style="font-size:0.85rem; color:#2e7d32; margin-top:6px; line-height:1.6;">
                    ✅ Summary KPIs · ✅ AFM Performance<br>
                    ✅ Andons by Type · ✅ Leaderboard · ✅ Raw Data
                </div>
            </div>""", unsafe_allow_html=True)
            with st.spinner("Generating Daily Excel…"):
                daily_bytes = report_builder.build_daily_report(fdf, uploaded_files, within_threshold)
            st.download_button("⬇️ Download Daily Excel", daily_bytes,
                               "LCY3_Daily_Report.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with d2:
            st.markdown("""
            <div style="background:#e8f5e9; border-left:5px solid #2e7d32; border-radius:8px;
                        padding:16px 20px; margin-bottom:12px;">
                <div style="font-size:1.1rem; font-weight:700; color:#1b5e20;">📄 Daily PDF Report</div>
                <div style="font-size:0.85rem; color:#2e7d32; margin-top:6px; line-height:1.6;">
                    ✅ Cover page · ✅ KPI summary<br>
                    ✅ Leaderboard · ✅ Andon Types · ✅ Flagged resolvers
                </div>
            </div>""", unsafe_allow_html=True)
            with st.spinner("Generating Daily PDF…"):
                try:
                    import pdf_report
                    daily_pdf = pdf_report.build_pdf_daily(fdf, uploaded_files, within_threshold)
                    st.download_button("⬇️ Download Daily PDF", daily_pdf,
                                       "LCY3_Daily_Report.pdf", "application/pdf",
                                       use_container_width=True)
                except Exception as e:
                    st.error(f"PDF generation failed: {e}")

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown('<div class="sec-title">📆 Weekly Reports</div>', unsafe_allow_html=True)
        w1, w2 = st.columns(2)

        with w1:
            st.markdown("""
            <div style="background:#e3f2fd; border-left:5px solid #1976d2; border-radius:8px;
                        padding:16px 20px; margin-bottom:12px;">
                <div style="font-size:1.1rem; font-weight:700; color:#0d47a1;">📊 Weekly Excel Report</div>
                <div style="font-size:0.85rem; color:#1565c0; margin-top:6px; line-height:1.6;">
                    ✅ Weekly KPIs · ✅ Andon Type x Week<br>
                    ✅ AFM x Week · ✅ System vs Non-System
                </div>
            </div>""", unsafe_allow_html=True)
            with st.spinner("Generating Weekly Excel…"):
                weekly_bytes = report_builder.build_weekly_report(fdf, uploaded_files, within_threshold)
            st.download_button("⬇️ Download Weekly Excel", weekly_bytes,
                               "LCY3_Weekly_Report.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with w2:
            st.markdown("""
            <div style="background:#e3f2fd; border-left:5px solid #1565c0; border-radius:8px;
                        padding:16px 20px; margin-bottom:12px;">
                <div style="font-size:1.1rem; font-weight:700; color:#0d47a1;">📄 Weekly PDF Report</div>
                <div style="font-size:0.85rem; color:#1565c0; margin-top:6px; line-height:1.6;">
                    ✅ Cover page · ✅ Weekly KPIs<br>
                    ✅ Week breakdown · ✅ System vs Non-System
                </div>
            </div>""", unsafe_allow_html=True)
            with st.spinner("Generating Weekly PDF…"):
                try:
                    import pdf_report
                    weekly_pdf = pdf_report.build_pdf_weekly(fdf, uploaded_files, within_threshold)
                    st.download_button("⬇️ Download Weekly PDF", weekly_pdf,
                                       "LCY3_Weekly_Report.pdf", "application/pdf",
                                       use_container_width=True)
                except Exception as e:
                    st.error(f"PDF generation failed: {e}")

        st.markdown("<br>", unsafe_allow_html=True)
        st.info("💡 Reports reflect your current filter selections. Adjust filters on any tab, then download.")

    # ── Tab: History ──────────────────────────────────────────────────────────
    with tab["📂 History"]:
        st.markdown('<div class="sec-title">Upload History</div>', unsafe_allow_html=True)
        hist_records = history_db.get_history(50)
        if not hist_records:
            st.info("No upload history yet. Upload a file to start building your history.")
        else:
            summary_rows = []
            for r in hist_records:
                summary_rows.append({
                    "File Name":    r["file_name"],
                    "Upload Date":  r.get("upload_date", r["upload_ts"][:10] if r["upload_ts"] else ""),
                    "Week(s)":      ", ".join([f"Wk {w}" for w in r["week_numbers"]]) or "—",
                    "Andons":       r["total_andons"],
                    "Date Range":   f"{r['date_min']} → {r['date_max']}",
                    "Daily Data":   "✅" if r.get("has_daily") else "—",
                    "Weekly Data":  "✅" if r.get("has_weekly") else "—",
                })
            st.markdown("#### All Saved Datasets")
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

            st.markdown("<br>", unsafe_allow_html=True)

            fname_options = [r["file_name"] for r in hist_records]
            sel_hist = st.selectbox("Inspect dataset", fname_options, key="hist_sel")
            sel_rec  = next(r for r in hist_records if r["file_name"] == sel_hist)
            sel_hash = sel_rec["file_hash"]

            h_daily, h_weekly, h_proc = st.tabs(["📅 Daily", "📆 Weekly", "🗂 Processed"])

            with h_daily:
                daily_df = history_db.load_daily(sel_hash)
                if daily_df is not None and not daily_df.empty:
                    st.markdown(f"**{len(daily_df)} day(s)** of data for **{sel_hist}**")
                    fmt = {}
                    if "Avg_Resolve_Min" in daily_df.columns:
                        fmt["Avg_Resolve_Min"] = "{:.2f}"
                    st.dataframe(daily_df.style.format(fmt), use_container_width=True, hide_index=True)
                    st.download_button("⬇️ Download daily CSV", daily_df.to_csv(index=False).encode(),
                                       f"{sel_hist}_daily.csv", "text/csv", use_container_width=True)
                else:
                    st.info("No daily breakdown available for this file.")

            with h_weekly:
                weekly_df = history_db.load_weekly(sel_hash)
                if weekly_df is not None and not weekly_df.empty:
                    st.markdown(f"**{len(weekly_df)} week(s)** of data for **{sel_hist}**")
                    fmt = {}
                    if "Avg_Resolve_Min" in weekly_df.columns:
                        fmt["Avg_Resolve_Min"] = "{:.2f}"
                    st.dataframe(weekly_df.style.format(fmt), use_container_width=True, hide_index=True)
                    st.download_button("⬇️ Download weekly CSV", weekly_df.to_csv(index=False).encode(),
                                       f"{sel_hist}_weekly.csv", "text/csv", use_container_width=True)
                else:
                    st.info("No weekly breakdown available for this file.")

            with h_proc:
                proc_df = history_db.load_dataframe(sel_hash)
                if proc_df is not None and not proc_df.empty:
                    st.markdown(f"**{len(proc_df):,} rows** · **{proc_df.shape[1]} columns**")
                    st.dataframe(proc_df, use_container_width=True, hide_index=True)
                    st.download_button("⬇️ Download processed CSV", proc_df.to_csv(index=False).encode(),
                                       f"{sel_hist}_processed.csv", "text/csv", use_container_width=True)
                else:
                    st.info("No processed data available for this file.")

else:
    st.markdown(f"""
    <div style="text-align:center; padding:4rem 2rem;
                background: {"linear-gradient(135deg, #1a1d27 0%, #1e2235 100%)" if DM else "linear-gradient(135deg, #f0f4ff 0%, #e8eaf6 100%)"};
                border-radius:16px; margin-top:1.5rem;
                border: 2px dashed {"rgba(121,134,203,0.35)" if DM else "rgba(57,73,171,0.25)"};
                box-shadow: {"0 4px 32px rgba(0,0,0,0.4)" if DM else "0 4px 20px rgba(57,73,171,0.1)"};
                animation: fadeUp 0.5s cubic-bezier(0.22,1,0.36,1);">
        <div style="font-size:3.5rem;margin-bottom:1rem;filter:drop-shadow(0 0 12px rgba(121,134,203,0.4));">📊</div>
        <h2 style="color:{"#c5cae9" if DM else "#1a237e"}; margin-bottom:0.5rem; font-size:1.6rem; font-weight:900; letter-spacing:-0.02em;">
            Welcome to the LCY3 AFM Dashboard
        </h2>
        <p style="color:{"#8892b0" if DM else "#555"}; font-size:1rem; margin-top:0.5rem;">
            Upload a <strong style="color:{"#7986cb" if DM else "#3949ab"};">JSON or CSV file</strong> above to explore your Andon data
        </p>
        <div style="margin-top:1.5rem; display:inline-flex; flex-direction:column; align-items:center; gap:0.5rem;">
            <div style="background:{"rgba(57,73,171,0.15)" if DM else "rgba(57,73,171,0.08)"}; border:1px solid {"rgba(121,134,203,0.25)" if DM else "rgba(57,73,171,0.15)"};
                         border-radius:10px; padding:0.7rem 1.5rem; font-size:0.82rem; color:{"#8892b0" if DM else "#555"}; text-align:left;">
                <div style="font-weight:700; color:{"#7986cb" if DM else "#3949ab"}; margin-bottom:4px;">✅ Required columns</div>
                Status &nbsp;·&nbsp; Resolver &nbsp;·&nbsp; Andon Type &nbsp;·&nbsp; Dwell Time (hh:mm:ss) &nbsp;·&nbsp; Time Created
            </div>
            <div style="background:{"rgba(57,73,171,0.08)" if DM else "rgba(57,73,171,0.04)"}; border:1px solid {"rgba(121,134,203,0.15)" if DM else "rgba(57,73,171,0.1)"};
                         border-radius:10px; padding:0.7rem 1.5rem; font-size:0.82rem; color:{"#8892b0" if DM else "#777"}; text-align:left;">
                <div style="font-weight:700; color:{"#5c6bc0" if DM else "#5c6bc0"}; margin-bottom:4px;">⚡ Optional columns (unlock more tabs)</div>
                Equipment Type &nbsp;·&nbsp; Zone &nbsp;·&nbsp; Shift &nbsp;·&nbsp; Blocking &nbsp;·&nbsp; Equipment ID &nbsp;·&nbsp; Creator &nbsp;·&nbsp; Time Resolved
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
